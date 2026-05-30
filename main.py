# ================= IMPORTS & KONFIGURASI =================
import sys
import os
import pandas as pd
from datetime import datetime
import numpy as np
# PyQt5 imports
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QFont, QColor, QPalette, QIcon
from PyQt5.QtWidgets import QShortcut

# Excel libraries
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill, Border, Side
import openpyxl

# Coba import LPDProcessor
try:
    from LPDProcessor import LPDProcessor
    LPD_AVAILABLE = True
    print("✓ LPDProcessor tersedia")
except ImportError as e:
    print(f"⚠ LPDProcessor tidak tersedia: {e}")
    LPD_AVAILABLE = False
    
    # Buat class dummy jika tidak tersedia
    class LPDProcessor:
        def process_excel_to_template(self, file_path):
            print("LPDProcessor tidak tersedia - menggunakan fallback")
            return pd.DataFrame()

# Mapping cabang
CABANG_KODE_MAP = {
    "UTM": "SURABAYA",
    "GBA": "BALI",
    "GTK": "MALANG",
    "GMT": "MOJOKERTO",
    "GPO": "PONOROGO",
    "GLB": "LUMAJANG",
    "GPB": "TAPAL KUDA",
    "GSM": "SEMARANG",
    "GJ": "JOGJA",
    "GDS": "SOLO",
    "GBD": "BANDUNG",
    "GJK": "JAKARTA",
    "GPN": "PEKANBARU",
    "GPD": "PADANG"
}


class ExcelReportProcessor(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📊 Excel Report Processor - Konsinyasi & Penjualan")
        self.setGeometry(50, 50, 1600, 900)

        # View manager
        self.konsinyasi_df = None  # Data konsinyasi
        self.penjualan_df = None   # Data penjualan
        self.konsinyasi_file = None
        self.penjualan_file = None
        self.current_view_mode = 0  # 0=Konsinyasi, 1=Penjualan, 2=Gabungan

        self.konsinyasi_processed = None  # Hasil processing konsinyasi
        self.penjualan_processed = None   # Hasil processing penjualan
        self.table_gabungan = None
        
        # SKU Management
        self.konsinyasi_skus = []  # SKU yang dideteksi dari data konsinyasi
        self.penjualan_skus = []   # SKU yang dideteksi dari data penjualan
        self.all_dynamic_skus = [] # Semua SKU dinamis yang terdeteksi

        # Default SKU (hardcoded)
        self.default_skus = [
            'TO', 'TB', 'TBL', 'TS', 'TOS4', 'TBS', 'TLS', 'TWS',
            'TOSB', 'TWSB', 
            'FK', 'FUK', 'FKS', 'PCB', 'PCC', 'PCK', 'PCP', 'KNT', 'KNC', 
            'KNW', 'YPB', 'YPC', 'YPS', 'PNN', 'PNT', 'ORI', 'DB', 'TA'
        ]
        
        # Gabungkan default + dynamic
        self.all_skus = self.default_skus.copy()
        
        # Setup UI
        self.setup_ui()
        
        # Setup menu dan shortcut
        self.setup_menu()
        self.setup_shortcuts()

        self.hide_old_elements()
        self.clear_all_tables()

    def clear_all_tables(self):
        """Clear all table widgets"""
        tables = [
            self.table_konsinyasi,
            self.table_penjualan,
            self.table_gabungan,  # Ganti 3 tabel dengan 1 tabel
        ]
        
        for table in tables:
            if table is not None:
                table.clear()
                table.setRowCount(0)
                table.setColumnCount(0)

    def clear_all_data(self):
        """Clear all loaded data"""
        reply = QMessageBox.question(self, "Konfirmasi",
            "Apakah Anda yakin ingin menghapus semua data?\n"
            "Data konsinyasi dan penjualan akan hilang.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            # Data baru saja
            self.konsinyasi_df = None
            self.penjualan_df = None
            self.konsinyasi_file = None
            self.penjualan_file = None
            self.konsinyasi_processed = None
            self.penjualan_processed = None
            
            # Reset UI semua tabel
            self.clear_all_tables()
            
            # Reset file labels
            self.kons_file_label.setText("Belum ada file dipilih")
            self.kons_file_label.setStyleSheet("color:#666; font-style:italic;")
            
            self.penj_file_label.setText("Belum ada file dipilih")
            self.penj_file_label.setStyleSheet("color:#666; font-style:italic;")
            
            # Reset stats label
            self.stats_label.setText("📊 Pilih file dan view yang diinginkan")
            
            # Reset info labels gabungan
            self.kons_info_label.setText("📋 KONSINYASI: 0 baris")
            self.penj_info_label.setText("💰 PENJUALAN: 0 baris")
            
            # Reset buttons
            self.kons_process_btn.setEnabled(False)
            self.penj_process_btn.setEnabled(False)
            self.export_btn.setEnabled(False)
            
            # Kembali ke view konsinyasi
            self.switch_view_mode(0)
            
            self.status_bar.showMessage("Semua data telah dihapus")
                
    def hide_old_elements(self):
        """Sembunyikan elemen sistem lama"""
        # Hanya update mode label
        self.mode_info_label.setText("Mode: VIEW 3-JENIS")
            
    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(12)

        # ================= HEADER =================
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(15, 10, 15, 10)

        self.title_label = QLabel("📊 Excel Report Processor")
        self.title_label.setFont(QFont("Arial", 18, QFont.Bold))
        self.title_label.setStyleSheet("""
            color: white;
            padding: 12px 20px;
            border-radius: 8px;
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #3498db, stop:1 #9b59b6);
        """)

        right_header = QVBoxLayout()
        right_header.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.subtitle_label = QLabel("Konsinyasi & Penjualan Report Viewer")
        self.subtitle_label.setStyleSheet("color:#7f8c8d; font-style:italic;")
        self.mode_info_label = QLabel("Mode: DATA INPUT")
        self.mode_info_label.setAlignment(Qt.AlignCenter)
        self.mode_info_label.setStyleSheet("""
            background:#3498db;
            color:white;
            padding:6px 14px;
            border-radius:12px;
            font-weight:bold;
        """)

        right_header.addWidget(self.subtitle_label)
        right_header.addWidget(self.mode_info_label)
        header_layout.addWidget(self.title_label, 1)
        header_layout.addLayout(right_header)
        main_layout.addWidget(header_widget)

        # ================= VIEW SELECTOR =================
        view_panel = QFrame()
        view_panel.setFrameShape(QFrame.StyledPanel)
        view_panel.setStyleSheet("""
            QFrame { 
                background:#f0f8ff; 
                border:2px solid #4682b4; 
                border-radius:8px; 
                padding:10px;
            }
        """)
        
        view_layout = QHBoxLayout(view_panel)
        view_layout.setContentsMargins(10, 10, 10, 10)
        
        self.view_konsinyasi_btn = QPushButton("📋 VIEW KONSINYASI")
        self.view_konsinyasi_btn.setCheckable(True)
        self.view_konsinyasi_btn.setChecked(True)
        self.view_konsinyasi_btn.clicked.connect(lambda: self.switch_view_mode(0))
        self.view_konsinyasi_btn.setStyleSheet(self.get_view_button_style(True))
        
        self.view_penjualan_btn = QPushButton("💰 VIEW PENJUALAN")
        self.view_penjualan_btn.setCheckable(True)
        self.view_penjualan_btn.clicked.connect(lambda: self.switch_view_mode(1))
        self.view_penjualan_btn.setStyleSheet(self.get_view_button_style(False))
        
        self.view_gabungan_btn = QPushButton("👁️ VIEW GABUNGAN")
        self.view_gabungan_btn.setCheckable(True)
        self.view_gabungan_btn.clicked.connect(lambda: self.switch_view_mode(2))
        self.view_gabungan_btn.setStyleSheet(self.get_view_button_style(False))
        
        view_layout.addWidget(QLabel("Pilih View:"))
        view_layout.addWidget(self.view_konsinyasi_btn)
        view_layout.addWidget(self.view_penjualan_btn)
        view_layout.addWidget(self.view_gabungan_btn)
        view_layout.addStretch()
        
        main_layout.addWidget(view_panel)
        
        # ================= FILE INPUT PANELS =================
        self.file_input_panel = QWidget()
        file_input_layout = QVBoxLayout(self.file_input_panel)
        
        # Panel Konsinyasi
        kons_panel = self.create_file_input_panel("KONSINYASI", "#e8f4f8")
        self.kons_file_label = kons_panel.findChild(QLabel, "file_label")
        self.kons_browse_btn = kons_panel.findChild(QPushButton, "browse_btn")
        self.kons_process_btn = kons_panel.findChild(QPushButton, "process_btn")
        self.kons_browse_btn.clicked.connect(lambda: self.browse_file("konsinyasi"))
        self.kons_process_btn.clicked.connect(lambda: self.process_file("konsinyasi"))
        
        # Panel Penjualan
        penj_panel = self.create_file_input_panel("PENJUALAN", "#f8e8f4")
        self.penj_file_label = penj_panel.findChild(QLabel, "file_label")
        self.penj_browse_btn = penj_panel.findChild(QPushButton, "browse_btn")
        self.penj_process_btn = penj_panel.findChild(QPushButton, "process_btn")
        self.penj_browse_btn.clicked.connect(lambda: self.browse_file("penjualan"))
        self.penj_process_btn.clicked.connect(lambda: self.process_file("penjualan"))
        
        file_input_layout.addWidget(kons_panel)
        file_input_layout.addWidget(penj_panel)
        
        main_layout.addWidget(self.file_input_panel)
        
        # ================= INFO PANEL =================
        info_panel = QFrame()
        info_panel.setFrameShape(QFrame.StyledPanel)
        info_panel.setStyleSheet("""
            QFrame { 
                background:#fff8e1; 
                border:1px solid #ffd54f; 
                border-radius:5px; 
                padding:8px;
            }
        """)
        info_layout = QHBoxLayout(info_panel)
        
        self.stats_label = QLabel("📊 Pilih file dan view yang diinginkan")
        self.stats_label.setStyleSheet("color:#555; font-weight:bold;")
        info_layout.addWidget(self.stats_label)
        
        # Tambah space dan export button
        info_layout.addStretch()
        
        self.export_btn = QPushButton("💾 Export Laporan")
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_data)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                border: 2px solid #219653;
                border-radius: 5px;
                padding: 8px 15px;
            }
            QPushButton:hover {
                background-color: #219653;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
                border-color: #95a5a6;
            }
        """)
        info_layout.addWidget(self.export_btn)
        
        main_layout.addWidget(info_panel)
        
        # ================= TABLE AREA =================
        self.table_area = QStackedWidget()
        
        # Tabel Konsinyasi - Index 0
        konsinyasi_widget = QWidget()
        konsinyasi_layout = QVBoxLayout(konsinyasi_widget)
        konsinyasi_layout.setContentsMargins(0, 0, 0, 0)
        
        kons_header = QLabel("📋 DATA KONSINYASI")
        kons_header.setFont(QFont("Arial", 12, QFont.Bold))
        kons_header.setStyleSheet("color:#2c3e50; margin:5px;")
        konsinyasi_layout.addWidget(kons_header)
        
        self.table_konsinyasi = QTableWidget()
        self.table_konsinyasi.setAlternatingRowColors(True)
        self.table_konsinyasi.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
                font-size: 11px;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                padding: 4px;
                border: 1px solid #2980b9;
            }
        """)
        konsinyasi_layout.addWidget(self.table_konsinyasi)
        self.table_area.addWidget(konsinyasi_widget)
        
        # Tabel Penjualan - Index 1
        penjualan_widget = QWidget()
        penjualan_layout = QVBoxLayout(penjualan_widget)
        penjualan_layout.setContentsMargins(0, 0, 0, 0)
        
        penj_header = QLabel("💰 DATA PENJUALAN")
        penj_header.setFont(QFont("Arial", 12, QFont.Bold))
        penj_header.setStyleSheet("color:#2c3e50; margin:5px;")
        penjualan_layout.addWidget(penj_header)
        
        self.table_penjualan = QTableWidget()
        self.table_penjualan.setAlternatingRowColors(True)
        self.table_penjualan.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
                font-size: 11px;
            }
            QHeaderView::section {
                background-color: #9b59b6;
                color: white;
                font-weight: bold;
                padding: 4px;
                border: 1px solid #8e44ad;
            }
        """)
        penjualan_layout.addWidget(self.table_penjualan)
        self.table_area.addWidget(penjualan_widget)

    # ================= WIDGET GABUNGAN - Index 2 =================
        gabungan_widget = QWidget()
        gabungan_layout = QVBoxLayout(gabungan_widget)
        gabungan_layout.setContentsMargins(10, 10, 10, 10)

        # Header untuk gabungan
        gabungan_header = QLabel("👁️ DATA GABUNGAN KONSINYASI & PENJUALAN")
        gabungan_header.setFont(QFont("Arial", 16, QFont.Bold))
        gabungan_header.setStyleSheet("""
            color: #2C3E50;
            background: linear-gradient(to right, #3498db, #9b59b6);
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 15px;
            text-align: center;
        """)
        gabungan_header.setFixedHeight(60)
        gabungan_layout.addWidget(gabungan_header)

        # Panel info untuk statistik
        info_frame = QFrame()
        info_frame.setFrameShape(QFrame.StyledPanel)
        info_frame.setStyleSheet("""
            QFrame {
                background: #f8f9fa;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                padding: 12px;
                margin-bottom: 15px;
            }
        """)
        info_layout = QHBoxLayout(info_frame)

        # Info konsinyasi
        self.kons_info_label = QLabel("📋 KONSINYASI: 0 baris")
        self.kons_info_label.setStyleSheet("""
            color: #2980b9; 
            font-weight: bold; 
            font-size: 13px;
            padding: 8px 15px;
            background: #e8f4f8;
            border-radius: 5px;
            border: 1px solid #3498db;
        """)
        info_layout.addWidget(self.kons_info_label)

        # Spacer
        info_layout.addSpacing(30)

        # Info penjualan
        self.penj_info_label = QLabel("💰 PENJUALAN: 0 baris")
        self.penj_info_label.setStyleSheet("""
            color: #8e44ad; 
            font-weight: bold; 
            font-size: 13px;
            padding: 8px 15px;
            background: #f8e8f4;
            border-radius: 5px;
            border: 1px solid #9b59b6;
        """)
        info_layout.addWidget(self.penj_info_label)

        info_layout.addStretch()
        gabungan_layout.addWidget(info_frame)

        # ================= TABEL TUNGGAL GABUNGAN =================
        # Hanya gunakan 1 tabel untuk menampung semua data
        self.table_gabungan = QTableWidget()
        self.table_gabungan.setAlternatingRowColors(True)
        self.table_gabungan.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
                font-size: 11px;
                selection-background-color: #3498db;
                background: white;
            }
            QHeaderView::section {
                background-color: #2C3E50;
                color: white;
                font-weight: bold;
                padding: 8px;
                border: 1px solid #34495e;
                font-size: 11px;
            }
            QTableWidget::item {
                padding: 4px;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)

        gabungan_layout.addWidget(self.table_gabungan)
        self.table_area.addWidget(gabungan_widget)
        
        # Tambahkan table area ke layout utama
        main_layout.addWidget(self.table_area, 1)
        
        # ================= PROGRESS BAR =================
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setMaximumHeight(18)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 5px;
            }
        """)
        main_layout.addWidget(self.progress_bar)
        
        # ================= STATUS BAR =================
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        
        # Inisialisasi tabel
        self.clear_all_tables()
    
    def calculate_sku_statistics_for_df(self, df, jenis, display=True):
        """Hitung statistik SKU untuk DataFrame individual (konsinyasi/penjualan)"""
        try:
            if df is None or df.empty:
                if display:
                    print(f"⚠ Data {jenis} kosong untuk statistik SKU")
                return {
                    'total_sku_ditemukan': 0,
                    'sku_masuk': 0,
                    'sku_tidak_masuk': 0,
                    'persen_masuk': 0,
                    'persen_tidak_masuk': 0,
                    'total_pelanggan': 0,
                    'sku_masuk_list': [],
                    'sku_tidak_masuk_list': []
                }
            
            if display:
                print(f"\n📊 PERHITUNGAN STATISTIK SKU UNTUK {jenis}:")
                print(f"{'-'*50}")
            
            # Dapatkan kolom SKU dari DataFrame
            sku_columns = self.get_sku_columns_from_df(df)
            if display:
                print(f"✓ Kolom SKU terdeteksi: {len(sku_columns)} kolom")
            
            # Hitung statistik per SKU
            sku_statistics = {}
            
            for sku in sku_columns:
                if sku in df.columns:
                    # PERBAIKAN: Konversi ke numeric dulu
                    try:
                        df_sku_numeric = pd.to_numeric(df[sku], errors='coerce')
                        # Hitung berapa pelanggan yang membeli SKU ini
                        pelanggan_dengan_sku = (df_sku_numeric > 0).sum()
                    except:
                        # Jika error, gunakan 0
                        pelanggan_dengan_sku = 0
                    
                    total_pelanggan = len(df)
                    
                    sku_statistics[sku] = {
                        'pelanggan_membeli': pelanggan_dengan_sku,
                        'total_pelanggan': total_pelanggan,
                        'persentase': (pelanggan_dengan_sku / total_pelanggan * 100) if total_pelanggan > 0 else 0
                    }
            
            # Filter SKU yang masuk dan tidak masuk
            sku_masuk = []
            sku_tidak_masuk = []
            
            for sku, stats in sku_statistics.items():
                if stats['pelanggan_membeli'] > 0:
                    sku_masuk.append({
                        'sku': sku,
                        'pelanggan_membeli': stats['pelanggan_membeli'],
                        'persentase': stats['persentase']
                    })
                else:
                    sku_tidak_masuk.append(sku)
            
            # Urutkan SKU masuk berdasarkan persentase (desc)
            sku_masuk_sorted = sorted(sku_masuk, key=lambda x: x['persentase'], reverse=True)
            
            # Hitung total
            total_sku_ditemukan = len(sku_columns)
            sku_masuk_count = len(sku_masuk)
            sku_tidak_masuk_count = len(sku_tidak_masuk)
            total_pelanggan = len(df)
            
            # Hitung persentase
            persen_masuk = (sku_masuk_count / total_sku_ditemukan * 100) if total_sku_ditemukan > 0 else 0
            persen_tidak_masuk = (sku_tidak_masuk_count / total_sku_ditemukan * 100) if total_sku_ditemukan > 0 else 0
            
            if display:
                print(f"📈 HASIL STATISTIK:")
                print(f"  • Total pelanggan: {total_pelanggan}")
                print(f"  • Total SKU ditemukan: {total_sku_ditemukan}")
                print(f"  • SKU masuk: {sku_masuk_count} ({persen_masuk:.1f}%)")
                print(f"  • SKU tidak masuk: {sku_tidak_masuk_count} ({persen_tidak_masuk:.1f}%)")
                
                if sku_masuk_sorted:
                    print(f"\n🏆 TOP 5 SKU TERPOPULER:")
                    for i, item in enumerate(sku_masuk_sorted[:5]):
                        print(f"  {i+1}. {item['sku']}: {item['pelanggan_membeli']}/{total_pelanggan} "
                            f"pelanggan ({item['persentase']:.1f}%)")
                
                if sku_tidak_masuk:
                    print(f"\n⚠ SKU YANG TIDAK TERJUAL:")
                    # Tampilkan maksimal 10 SKU
                    display_skus = sku_tidak_masuk[:10]
                    print(f"  {', '.join(display_skus)}")
                    if len(sku_tidak_masuk) > 10:
                        print(f"  ... dan {len(sku_tidak_masuk) - 10} SKU lainnya")
            
            return {
                'total_sku_ditemukan': total_sku_ditemukan,
                'sku_masuk': sku_masuk_count,
                'sku_tidak_masuk': sku_tidak_masuk_count,
                'persen_masuk': persen_masuk,
                'persen_tidak_masuk': persen_tidak_masuk,
                'total_pelanggan': total_pelanggan,
                'sku_masuk_list': [item['sku'] for item in sku_masuk_sorted],
                'sku_tidak_masuk_list': sku_tidak_masuk,
                'top_skus': sku_masuk_sorted[:5] if sku_masuk_sorted else []
            }
            
        except Exception as e:
            print(f"❌ Error calculating SKU statistics for {jenis}: {e}")
            import traceback
            traceback.print_exc()
            return {
                'total_sku_ditemukan': 0,
                'sku_masuk': 0,
                'sku_tidak_masuk': 0,
                'persen_masuk': 0,
                'persen_tidak_masuk': 0,
                'total_pelanggan': 0,
                'sku_masuk_list': [],
                'sku_tidak_masuk_list': []
            }
    
    

    def _convert_column_to_numeric(self, series):
        """Konversi series ke numeric dengan handling error"""
        try:
            return pd.to_numeric(series, errors='coerce')
        except:
            # Jika gagal, coba ekstrak angka dari string
            def extract_number(x):
                if pd.isna(x):
                    return np.nan
                if isinstance(x, (int, float)):
                    return float(x)
                if isinstance(x, str):
                    # Cari angka dalam string
                    import re
                    numbers = re.findall(r'\d+\.?\d*', str(x))
                    if numbers:
                        try:
                            return float(numbers[0])
                        except:
                            return np.nan
                return np.nan
            
            return series.apply(extract_number)
        
    def _create_total_row_with_sku_stats(self, system_name, description, source_row, columns, has_source, jenis):
        """Buat satu baris total dengan kolom SKU statistik - DIPERBAIKI"""
        row = []
        
        # 1. Tambahkan NAMA PELANGGAN (deskripsi)
        row.append(description)
        
        # 2. Tambahkan SISTEM
        row.append(system_name)
        
        # 3. Tambahkan TOTAL PENJUALAN dari source_row
        total_penjualan = 0
        if has_source and source_row is not None:
            # Cari nilai TOTAL PENJUALAN di semua kolom
            for col_name, value in source_row.items():
                if 'TOTAL' in str(col_name).upper() and 'PENJUALAN' in str(col_name).upper():
                    total_penjualan = self._extract_number_from_value(value)
                    break
        
        row.append(total_penjualan)
        
        # 4. Dapatkan DataFrame yang sesuai untuk perhitungan statistik
        if jenis == "konsinyasi":
            df_for_stats = self.konsinyasi_processed
        else:
            df_for_stats = self.penjualan_processed
        
        # 5. Urutkan SKU sesuai priority
        priority_skus = ['TWS', 'TWSB', 'YPB', 'YPC', 'YPS']
        
        # Urutkan semua SKU: priority dulu, lalu alfabet
        ordered_skus = []
        
        # Tambahkan priority SKU dulu
        for sku in priority_skus:
            if sku in self.all_skus:
                ordered_skus.append(sku)
        
        # Tambahkan SKU lainnya kecuali yang sudah ada
        for sku in self.all_skus:
            if sku not in ordered_skus:
                ordered_skus.append(sku)
        
        # 6. Tambahkan nilai untuk setiap SKU dari source_row
        sku_masuk_list = []
        sku_tidak_masuk_list = []
        
        for sku in ordered_skus:
            # Cari nilai di source_row
            sku_value = 0
            if has_source and source_row is not None:
                # Cari dengan case-insensitive
                for col_name, value in source_row.items():
                    if str(col_name).upper().strip() == sku.upper():
                        sku_value = self._extract_number_from_value(value)
                        break
            
            row.append(sku_value)
            
            # Catat untuk statistik awal (akan diperbarui nanti)
            if sku_value > 0:
                sku_masuk_list.append(sku)
            else:
                sku_tidak_masuk_list.append(sku)
        
        # 7. **PERBAIKAN: HITUNG STATISTIK YANG BENAR dari data lengkap**
        print(f"\n🔍 Menghitung statistik SKU untuk {system_name} dari data lengkap:")
        
        if df_for_stats is not None and not df_for_stats.empty:
            # Hitung ulang statistik dari data sebenarnya
            sku_masuk_list = []  # Reset dan hitung ulang
            sku_tidak_masuk_list = []  # Reset dan hitung ulang
            
            for sku in ordered_skus:
                if sku in df_for_stats.columns:
                    try:
                        # Konversi ke numeric
                        df_sku_numeric = pd.to_numeric(df_for_stats[sku], errors='coerce')
                        # Hitung berapa pelanggan yang membeli SKU ini (nilai > 0)
                        pelanggan_dengan_sku = (df_sku_numeric > 0).sum()
                        
                        if pelanggan_dengan_sku > 0:
                            sku_masuk_list.append(sku)
                        else:
                            sku_tidak_masuk_list.append(sku)
                            
                    except Exception as e:
                        sku_tidak_masuk_list.append(sku)
                else:
                    sku_tidak_masuk_list.append(sku)
            
            total_pelanggan = len(df_for_stats)
            print(f"  ✓ Data: {total_pelanggan} pelanggan")
            
        else:
            # Jika tidak ada data, gunakan perhitungan dari source_row saja
            print(f"  ⚠ Tidak ada data lengkap, gunakan perhitungan dari baris total")
        
        total_sku = len(ordered_skus)
        sku_masuk_count = len(sku_masuk_list)
        persen_masuk = (sku_masuk_count / total_sku * 100) if total_sku > 0 else 0
        persen_tidak_masuk = 100 - persen_masuk
        
        print(f"  ✓ {system_name}: {sku_masuk_count}/{total_sku} SKU terjual ({persen_masuk:.1f}%)")
        
        # 8. Format string untuk kolom statistik SKU
        sku_masuk_str = ", ".join(sku_masuk_list)  # TAMPILKAN SEMUA
        sku_tidak_str = ", ".join(sku_tidak_masuk_list)  # TAMPILKAN SEMUA
        
        # 9. Tambahkan kolom SKU statistik
        row.extend([
            sku_masuk_str,          # SKU YANG SUDAH MASUK
            sku_tidak_str,          # SKU YANG BELUM MASUK
            f"{sku_masuk_count}/{total_sku}",  # JUMLAH SKU MASUK
            f"{persen_masuk:.1f}%",  # %
            f"{len(sku_tidak_masuk_list)}/{total_sku}",  # JUMLAH SKU TIDAK MASUK
            f"{persen_tidak_masuk:.1f}%",  # %
            ""  # PERINGKAT SKU (kosong untuk total)
        ])
        
        return row


    def _is_sku_stat_column(self, column_name):
        """Cek apakah kolom adalah kolom statistik SKU"""
        col_str = str(column_name).upper().strip()
        
        stat_keywords = [
            'SKU YANG SUDAH MASUK',    # Format baru sesuai gambar
            'SKU YANG BELUM MASUK',    # Format baru sesuai gambar
            'JUMLAH SKU MASUK',        # Format baru sesuai gambar
            'JUMLAH SKU TIDAK MASUK',  # Format baru sesuai gambar
            'PERINGKAT SKU'
        ]
        
        return any(keyword in col_str for keyword in stat_keywords) or col_str == '%'

    def _is_pure_sku_column(self, column_name):
        """Cek apakah kolom adalah SKU murni (bukan peringkat/persentase)"""
        col_str = str(column_name).upper().strip()
        
        # Skip jika mengandung kata kunci peringkat/persentase
        exclude_keywords = ['PENCAPAIAN', 'PERSEN', '%', 'RANK', 'PERINGKAT', 'PROGRESS', 'PROSENTASE']
        if any(keyword in col_str for keyword in exclude_keywords):
            return False
        
        # Cek pattern SKU: 2-4 karakter, minimal 1 huruf, tidak semua angka
        if 2 <= len(col_str) <= 6:
            # Pattern: huruf semua atau kombinasi huruf-angka, tidak mengandung spasi/tanda khusus
            if col_str.isalpha() or (any(c.isalpha() for c in col_str) and not any(c.isspace() for c in str(column_name))):
                # Pastikan bukan kolom umum
                common_columns = ['DATE', 'TANGGAL', 'CABANG', 'BRANCH', 'KODE', 'ID', 'NO', 'JENIS']
                if col_str not in [c.upper() for c in common_columns]:
                    return True
        
        # Cek apakah termasuk dalam default SKU atau dynamic SKU
        if col_str in [sku.upper() for sku in self.all_skus]:
            return True
        
        return False

    def _extract_number_from_value(self, val):
        """Extract number dari berbagai format value"""
        if val is None:
            return 0
        
        if pd.isna(val):
            return 0
        
        # Jika sudah angka, langsung return
        if isinstance(val, (int, float)):
            return float(val)
        
        # Jika string, coba parse
        if isinstance(val, str):
            val_str = str(val).strip()
            
            # Format "X dari Y" (contoh: "36 dari 36", "1 dari 36")
            if 'dari' in val_str.lower():
                try:
                    # Ambil angka pertama sebelum "dari"
                    parts = val_str.split()
                    for part in parts:
                        # Cari angka
                        clean_part = ''.join(filter(str.isdigit, part))
                        if clean_part:
                            return float(clean_part)
                except:
                    pass
            
            # Format dengan koma, spasi, dll
            clean_val = val_str.replace(',', '').replace(' ', '').replace('Rp', '').strip()
            
            # Coba parse angka
            try:
                if clean_val and clean_val.replace('.', '').isdigit():
                    return float(clean_val)
                elif clean_val.isdigit():
                    return float(clean_val)
            except:
                pass
            
            # Coba ekstrak angka dari string apapun
            import re
            numbers = re.findall(r'\d+', clean_val)
            if numbers:
                try:
                    return float(numbers[0])
                except:
                    pass
        
        return 0
            
    def update_gabungan_view(self):
        """Update tampilan gabungan dengan semua data dalam 1 tabel - HEADER GABUNGAN"""
        try:
            if self.current_view_mode != 2:
                return
            
            print(f"\n{'='*60}")
            print(f"UPDATE GABUNGAN VIEW - DENGAN PERINGKAT SKU PER PELANGGAN")
            print(f"{'='*60}")
            
            # ================= 1. DAPATKAN HEADER GABUNGAN =================
            # A. Header dasar (selalu sama)
            base_headers = ['NAMA PELANGGAN', 'TOTAL PENJUALAN']
            
            # B. Kumpulkan SEMUA SKU dari kedua data
            all_unique_skus = set(self.default_skus)  # Mulai dari default
            
            # Tambahkan SKU dari konsinyasi jika ada
            if self.konsinyasi_processed is not None:
                kons_skus = self.get_sku_columns_from_df(self.konsinyasi_processed)
                for sku in kons_skus:
                    sku_str = str(sku).strip().upper()
                    if sku_str not in [s.upper() for s in all_unique_skus]:
                        all_unique_skus.add(sku)
            
            # Tambahkan SKU dari penjualan jika ada
            if self.penjualan_processed is not None:
                penj_skus = self.get_sku_columns_from_df(self.penjualan_processed)
                for sku in penj_skus:
                    sku_str = str(sku).strip().upper()
                    if sku_str not in [s.upper() for s in all_unique_skus]:
                        all_unique_skus.add(sku)
            
            # Tambahkan SKU dynamic yang sudah terdeteksi
            for sku in self.all_dynamic_skus:
                if sku.upper() not in [s.upper() for s in all_unique_skus]:
                    all_unique_skus.add(sku)
            
            # URUTKAN SKU: priority SKU dulu, lalu alfabet
            priority_skus = ['TWS', 'TWSB', 'YPB', 'YPC', 'YPS']
            
            # Konversi ke list untuk sorting
            all_unique_skus_list = list(all_unique_skus)
            
            # Fungsi untuk mendapatkan nilai priority
            def get_priority(sku):
                sku_str = str(sku).upper()
                if sku_str in priority_skus:
                    return priority_skus.index(sku_str)
                return 99  # Nilai tinggi untuk non-priority
            
            # Urutkan: priority dulu, lalu alfabet
            all_unique_skus_list.sort(key=lambda x: (get_priority(x), str(x).upper()))
            ordered_skus = all_unique_skus_list
            
            # C. Header statistik SKU (tetap sama)
            stat_headers = [
                'SKU YANG SUDAH MASUK', 'SKU YANG BELUM MASUK',
                'JUMLAH SKU MASUK', '%', 'JUMLAH SKU TIDAK MASUK', '%', 'PERINGKAT SKU'
            ]
            
            # Gabungkan semua header
            all_headers = ['JENIS'] + base_headers + ordered_skus + stat_headers
            
            print(f"✓ Header gabungan: {len(all_headers)} kolom")
            print(f"  - {len(base_headers)} kolom dasar")
            print(f"  - {len(ordered_skus)} kolom SKU (semua dari kedua data)")
            print(f"  - {len(stat_headers)} kolom statistik")
            print(f"  - Total SKU unik: {len(ordered_skus)}")
            print(f"  - Contoh SKU: {ordered_skus[:10]}...")
            
            # ================= 2. HITUNG PERINGKAT SKU UNTUK SEMUA PELANGGAN =================
            print(f"\n🏆 MENGHITUNG PERINGKAT SKU UNTUK SEMUA PELANGGAN:")
            
            # Hitung ranking untuk konsinyasi
            kons_ranking_dict = {}
            if self.konsinyasi_processed is not None and not self.konsinyasi_processed.empty:
                kons_ranking_dict = self.calculate_sku_ranking_for_all_customers(
                    self.konsinyasi_processed, 
                    ordered_skus, 
                    "KONSINYASI"
                )
                print(f"  ✓ Ranking konsinyasi: {len(kons_ranking_dict)} pelanggan")
            
            # Hitung ranking untuk penjualan
            penj_ranking_dict = {}
            if self.penjualan_processed is not None and not self.penjualan_processed.empty:
                penj_ranking_dict = self.calculate_sku_ranking_for_all_customers(
                    self.penjualan_processed, 
                    ordered_skus, 
                    "PENJUALAN"
                )
                print(f"  ✓ Ranking penjualan: {len(penj_ranking_dict)} pelanggan")
            
            # ================= 3. BUAT DATAFRAME GABUNGAN =================
            all_rows = []
            kons_rows = []  # Simpan sementara baris konsinyasi
            penj_rows = []  # Simpan sementara baris penjualan
            
            # A. Tambahkan data KONSINYASI jika ada
            if self.konsinyasi_processed is not None and not self.konsinyasi_processed.empty:
                kons_df = self.konsinyasi_processed.copy()
                
                # Update info label
                kons_stats = self.calculate_sku_statistics_for_df(
                    self.konsinyasi_processed, 
                    "KONSINYASI",
                    display=False
                )
                self.kons_info_label.setText(
                    f"📋 KONSINYASI: {len(kons_df)} baris | "
                    f"SKU: {kons_stats['sku_masuk']}/{kons_stats['total_sku_ditemukan']} "
                    f"({kons_stats['persen_masuk']:.1f}%)"
                )
                
                print(f"📋 Proses {len(kons_df)} baris konsinyasi dengan ranking SKU")
                
                # Konversi setiap baris konsinyasi ke format header gabungan
                row_count = 0
                for idx, row in kons_df.iterrows():
                    new_row = ['KONSINYASI']  # Kolom JENIS
                    
                    # 1. Tambahkan data dasar (NAMA PELANGGAN, SISTEM, TOTAL PENJUALAN)
                    for header in base_headers:
                        if header in kons_df.columns:
                            value = row[header]
                            # Format khusus untuk TOTAL PENJUALAN
                            if header == 'TOTAL PENJUALAN':
                                value = self._extract_number_from_value(value)
                            new_row.append(value)
                        else:
                            # Coba cari kolom yang mirip
                            found = False
                            for col in kons_df.columns:
                                if header.upper() in str(col).upper():
                                    value = row[col]
                                    if header == 'TOTAL PENJUALAN':
                                        value = self._extract_number_from_value(value)
                                    new_row.append(value)
                                    found = True
                                    break
                            if not found:
                                new_row.append("")  # Kosong jika tidak ada
                    
                    # 2. Tambahkan data untuk semua SKU
                    sku_masuk_list = []  # Untuk statistik per baris
                    sku_tidak_masuk_list = []  # Untuk statistik per baris
                    
                    for sku in ordered_skus:
                        if sku in kons_df.columns:
                            value = row[sku]
                            qty = self._extract_number_from_value(value)
                            new_row.append(qty)
                            
                            # Catat untuk statistik
                            if qty > 0:
                                sku_masuk_list.append(sku)
                            else:
                                sku_tidak_masuk_list.append(sku)
                        else:
                            # Coba cari dengan case-insensitive
                            found = False
                            for col in kons_df.columns:
                                if str(sku).upper() == str(col).upper():
                                    value = row[col]
                                    qty = self._extract_number_from_value(value)
                                    new_row.append(qty)
                                    
                                    # Catat untuk statistik
                                    if qty > 0:
                                        sku_masuk_list.append(sku)
                                    else:
                                        sku_tidak_masuk_list.append(sku)
                                    found = True
                                    break
                            if not found:
                                new_row.append(0)  # 0 untuk SKU yang tidak ada di data ini
                                sku_tidak_masuk_list.append(sku)
                    
                    # 3. HITUNG DAN TAMBAHKAN STATISTIK SKU untuk baris ini
                    sku_masuk_count = len(sku_masuk_list)
                    total_sku = len(ordered_skus)
                    persen_masuk = (sku_masuk_count / total_sku * 100) if total_sku > 0 else 0
                    persen_tidak_masuk = 100 - persen_masuk
                    
                    # Format data statistik - TAMPILKAN SEMUA TANPA TRUNCATION
                    sku_masuk_str = ", ".join(sku_masuk_list) if sku_masuk_list else "-"
                    sku_tidak_str = ", ".join(sku_tidak_masuk_list) if sku_tidak_masuk_list else "-"
                    
                    # Tambahkan kolom statistik SKU
                    new_row.append(sku_masuk_str)          # SKU YANG SUDAH MASUK
                    new_row.append(sku_tidak_str)          # SKU YANG BELUM MASUK
                    new_row.append(f"{sku_masuk_count}/{total_sku}")  # JUMLAH SKU MASUK
                    new_row.append(f"{persen_masuk:.1f}%")  # %
                    new_row.append(f"{len(sku_tidak_masuk_list)}/{total_sku}")  # JUMLAH SKU TIDAK MASUK
                    new_row.append(f"{persen_tidak_masuk:.1f}%")  # %
                    
                    # 4. HITUNG PERINGKAT SKU untuk baris ini - CEK DARI DATA YANG SUDAH ADA
                    ranking_str = ""

                    # Dapatkan nama pelanggan untuk lookup
                    nama_pelanggan = ""
                    for col in kons_df.columns:
                        if any(keyword in str(col).upper() for keyword in ['NAMA', 'PELANGGAN', 'CUSTOMER']):
                            nama_pelanggan = str(row[col]) if pd.notna(row[col]) else ""
                            break

                    # Cek apakah ini baris TOTAL - jika ya, kosongkan ranking
                    if 'TOTAL' in str(nama_pelanggan).upper():
                        ranking_str = ""  # Kosong untuk baris TOTAL
                        print(f"  ✓ Baris {row_count}: Baris TOTAL, ranking dikosongkan")
                    else:
                        # **PERBAIKAN UTAMA: CEK APAKAH DATA SUDAH PUNYA RANKING**
                        
                        # 1. Cek apakah di DataFrame sudah ada kolom 'PERINGKAT SKU'
                        if 'PERINGKAT SKU' in kons_df.columns:
                            ranking_value = row.get('PERINGKAT SKU')
                            if pd.notna(ranking_value):
                                ranking_str = str(ranking_value)
                                print(f"  ✓ Baris {row_count}: Ambil ranking langsung dari data ({ranking_str[:30]}...)")
                            else:
                                # Coba cari di mapping
                                if nama_pelanggan and nama_pelanggan.upper() in kons_ranking_map:
                                    ranking_str = kons_ranking_map[nama_pelanggan.upper()]
                                    print(f"  ✓ Baris {row_count}: Ambil ranking dari mapping ({ranking_str[:30]}...)")
                        else:
                            # Data tidak punya kolom ranking, hitung manual
                            sku_totals = []
                            for sku in ordered_skus:
                                if sku in kons_df.columns:
                                    value = row[sku]
                                    qty = self._extract_number_from_value(value)
                                    if qty > 0:
                                        sku_totals.append({'sku': sku, 'qty': qty})
                            
                            sku_totals_sorted = sorted(sku_totals, key=lambda x: x['qty'], reverse=True)
                            
                            if sku_totals_sorted:
                                # Ambil semua SKU, bukan hanya 3
                                all_skus_ranking = []
                                for i, item in enumerate(sku_totals_sorted):
                                    all_skus_ranking.append(f"{i+1}. {item['sku']} ({item['qty']:.0f} pcs)")
                                
                                # Batasi tampilan maksimal 5 SKU untuk readability
                                if len(all_skus_ranking) > 5:
                                    ranking_str = " | ".join(all_skus_ranking[:5]) + f" | ... ({len(all_skus_ranking)-5} SKU lainnya)"
                                else:
                                    ranking_str = " | ".join(all_skus_ranking)
                                
                                print(f"  ✓ Baris {row_count}: Hitung ranking baru ({len(sku_totals_sorted)} SKU)")
                            else:
                                ranking_str = "Tidak ada pembelian"
                                print(f"  ✓ Baris {row_count}: Tidak ada pembelian")
                    
                    # Tambahkan kolom PERINGKAT SKU dengan ranking yang sudah dihitung
                    new_row.append(ranking_str)  # PERINGKAT SKU diisi dengan ranking
                    
                    all_rows.append(new_row)
                    # Tambahkan baris ke list sementara konsinyasi
                    kons_rows.append({
                        'row_data': new_row,
                        'nama': nama_pelanggan,
                        'jenis': 'KONSINYASI',
                        'row_index': idx
                    })
                    
                    row_count += 1

                     # Progress indicator
                    if row_count % 50 == 0:
                        print(f"  - Diproses {row_count} baris konsinyasi...")
                
                print(f"✓ Selesai: {row_count} baris konsinyasi")

            else:
                print(f"📋 Data konsinyasi kosong")
                self.kons_info_label.setText("📋 KONSINYASI: 0 baris | SKU: 0/0 (0%)")
            
            # B. Tambahkan data PENJUALAN jika ada
            if self.penjualan_processed is not None and not self.penjualan_processed.empty:
                penj_df = self.penjualan_processed.copy()
                
                # Update info label
                penj_stats = self.calculate_sku_statistics_for_df(
                    self.penjualan_processed, 
                    "PENJUALAN",
                    display=False
                )
                self.penj_info_label.setText(
                    f"💰 PENJUALAN: {len(penj_df)} baris | "
                    f"SKU: {penj_stats['sku_masuk']}/{penj_stats['total_sku_ditemukan']} "
                    f"({penj_stats['persen_masuk']:.1f}%)"
                )
                
                print(f"💰 Proses {len(penj_df)} baris penjualan dengan ranking SKU")
                
                # Konversi setiap baris penjualan ke format header gabungan
                row_count = 0
                for idx, row in penj_df.iterrows():
                    new_row = ['PENJUALAN']  # Kolom JENIS
                    
                    # 1. Tambahkan data dasar (NAMA PELANGGAN, SISTEM, TOTAL PENJUALAN)
                    for header in base_headers:
                        if header in penj_df.columns:
                            value = row[header]
                            # Format khusus untuk TOTAL PENJUALAN
                            if header == 'TOTAL PENJUALAN':
                                value = self._extract_number_from_value(value)
                            new_row.append(value)
                        else:
                            # Coba cari kolom yang mirip
                            found = False
                            for col in penj_df.columns:
                                if header.upper() in str(col).upper():
                                    value = row[col]
                                    if header == 'TOTAL PENJUALAN':
                                        value = self._extract_number_from_value(value)
                                    new_row.append(value)
                                    found = True
                                    break
                            if not found:
                                new_row.append("")  # Kosong jika tidak ada
                    
                    # 2. Tambahkan data untuk semua SKU
                    sku_masuk_list = []  # Untuk statistik per baris
                    sku_tidak_masuk_list = []  # Untuk statistik per baris
                    
                    for sku in ordered_skus:
                        if sku in penj_df.columns:
                            value = row[sku]
                            qty = self._extract_number_from_value(value)
                            new_row.append(qty)
                            
                            # Catat untuk statistik
                            if qty > 0:
                                sku_masuk_list.append(sku)
                            else:
                                sku_tidak_masuk_list.append(sku)
                        else:
                            # Coba cari dengan case-insensitive
                            found = False
                            for col in penj_df.columns:
                                if str(sku).upper() == str(col).upper():
                                    value = row[col]
                                    qty = self._extract_number_from_value(value)
                                    new_row.append(qty)
                                    
                                    # Catat untuk statistik
                                    if qty > 0:
                                        sku_masuk_list.append(sku)
                                    else:
                                        sku_tidak_masuk_list.append(sku)
                                    found = True
                                    break
                            if not found:
                                new_row.append(0)  # 0 untuk SKU yang tidak ada di data ini
                                sku_tidak_masuk_list.append(sku)
                    
                    # 3. HITUNG DAN TAMBAHKAN STATISTIK SKU untuk baris ini
                    sku_masuk_count = len(sku_masuk_list)
                    total_sku = len(ordered_skus)
                    persen_masuk = (sku_masuk_count / total_sku * 100) if total_sku > 0 else 0
                    persen_tidak_masuk = 100 - persen_masuk
                    
                    # Format data statistik - TAMPILKAN SEMUA TANPA TRUNCATION
                    sku_masuk_str = ", ".join(sku_masuk_list) if sku_masuk_list else "-"
                    sku_tidak_str = ", ".join(sku_tidak_masuk_list) if sku_tidak_masuk_list else "-"
                    
                    # Tambahkan kolom statistik SKU
                    new_row.append(sku_masuk_str)          # SKU YANG SUDAH MASUK
                    new_row.append(sku_tidak_str)          # SKU YANG BELUM MASUK
                    new_row.append(f"{sku_masuk_count}/{total_sku}")  # JUMLAH SKU MASUK
                    new_row.append(f"{persen_masuk:.1f}%")  # %
                    new_row.append(f"{len(sku_tidak_masuk_list)}/{total_sku}")  # JUMLAH SKU TIDAK MASUK
                    new_row.append(f"{persen_tidak_masuk:.1f}%")  # %
                    
                    # 4. HITUNG PERINGKAT SKU untuk baris ini - CEK DARI DATA YANG SUDAH ADA
                    ranking_str = ""

                    # Dapatkan nama pelanggan untuk lookup
                    nama_pelanggan = ""
                    for col in penj_df.columns:
                        if any(keyword in str(col).upper() for keyword in ['NAMA', 'PELANGGAN', 'CUSTOMER']):
                            nama_pelanggan = str(row[col]) if pd.notna(row[col]) else ""
                            break

                    # Cek apakah ini baris TOTAL - jika ya, kosongkan ranking
                    if 'TOTAL' in str(nama_pelanggan).upper():
                        ranking_str = ""  # Kosong untuk baris TOTAL
                        print(f"  ✓ Baris {row_count}: Baris TOTAL, ranking dikosongkan")
                    else:
                        # **PERBAIKAN UTAMA: CEK APAKAH DATA SUDAH PUNYA RANKING**
                        
                        # 1. Cek apakah di DataFrame sudah ada kolom 'PERINGKAT SKU'
                        if 'PERINGKAT SKU' in penj_df.columns:
                            ranking_value = row.get('PERINGKAT SKU')
                            if pd.notna(ranking_value):
                                ranking_str = str(ranking_value)
                                print(f"  ✓ Baris {row_count}: Ambil ranking langsung dari data ({ranking_str[:30]}...)")
                            else:
                                # Coba cari di mapping
                                if nama_pelanggan and nama_pelanggan.upper() in penj_ranking_map:
                                    ranking_str = penj_ranking_map[nama_pelanggan.upper()]
                                    print(f"  ✓ Baris {row_count}: Ambil ranking dari mapping ({ranking_str[:30]}...)")
                        else:
                            # Data tidak punya kolom ranking, hitung manual
                            sku_totals = []
                            for sku in ordered_skus:
                                if sku in penj_df.columns:
                                    value = row[sku]
                                    qty = self._extract_number_from_value(value)
                                    if qty > 0:
                                        sku_totals.append({'sku': sku, 'qty': qty})
                            
                            sku_totals_sorted = sorted(sku_totals, key=lambda x: x['qty'], reverse=True)
                            
                            if sku_totals_sorted:
                                # Ambil semua SKU, bukan hanya 3
                                all_skus_ranking = []
                                for i, item in enumerate(sku_totals_sorted):
                                    all_skus_ranking.append(f"{i+1}. {item['sku']} ({item['qty']:.0f} pcs)")
                                
                                # Batasi tampilan maksimal 5 SKU untuk readability
                                if len(all_skus_ranking) > 5:
                                    ranking_str = " | ".join(all_skus_ranking[:5]) + f" | ... ({len(all_skus_ranking)-5} SKU lainnya)"
                                else:
                                    ranking_str = " | ".join(all_skus_ranking)
                                
                                print(f"  ✓ Baris {row_count}: Hitung ranking baru ({len(sku_totals_sorted)} SKU)")
                            else:
                                ranking_str = "Tidak ada pembelian"
                                print(f"  ✓ Baris {row_count}: Tidak ada pembelian")
                    
                    # Tambahkan kolom PERINGKAT SKU dengan ranking yang sudah dihitung
                    new_row.append(ranking_str)  # PERINGKAT SKU diisi dengan ranking
                    
                    all_rows.append(new_row)
                    # Tambahkan baris ke list sementara penjualan
                    penj_rows.append({
                        'row_data': new_row,
                        'nama': nama_pelanggan,
                        'jenis': 'PENJUALAN',
                        'row_index': idx
                    })
                    
                    row_count += 1
                    
                    # Progress indicator
                    if row_count % 50 == 0:
                        print(f"  - Diproses {row_count} baris penjualan...")
                
                print(f"✓ Selesai: {row_count} baris penjualan")
            
            else:
                print(f"💰 Data penjualan kosong")
                self.penj_info_label.setText("💰 PENJUALAN: 0 baris | SKU: 0/0 (0%)")
                
            # ================= 4. CARI DAN COPAS BARIS TOTAL PENCAPAIAN =================
            print(f"\n📊 Mencari baris TOTAL PENCAPAIAN untuk di-copas...")

            # Fungsi untuk mencari baris TOTAL PENCAPAIAN
            def find_total_pencapaian_row(df, jenis):
                if df is None or df.empty:
                    return None
                
                # Cari di semua kolom yang mengandung 'NAMA' atau 'PELANGGAN'
                nama_cols = [col for col in df.columns if any(keyword in str(col).upper() for keyword in ['NAMA', 'PELANGGAN', 'CUSTOMER'])]
                
                for col in nama_cols:
                    mask = df[col].astype(str).str.contains('TOTAL PENCAPAIAN', case=False, na=False)
                    if mask.any():
                        print(f"  ✓ TOTAL PENCAPAIAN ditemukan di {jenis} (kolom: {col})")
                        return df[mask].iloc[0].copy()
                
                # Jika tidak ditemukan, coba cari baris dengan kata "TOTAL" di kolom pertama
                first_col = df.columns[0]
                mask = df[first_col].astype(str).str.contains('TOTAL', case=False, na=False)
                if mask.any():
                    print(f"  ⚠ TOTAL (tanpa PENCAPAIAN) ditemukan di {jenis}")
                    return df[mask].iloc[0].copy()
                
                print(f"  ❌ Tidak ditemukan baris TOTAL di {jenis}")
                return None

            # Cari baris TOTAL PENCAPAIAN dari masing-masing data detail
            kons_total_detail = None
            penj_total_detail = None

            if self.konsinyasi_processed is not None:
                kons_total_detail = find_total_pencapaian_row(self.konsinyasi_processed, "KONSINYASI")

            if self.penjualan_processed is not None:
                penj_total_detail = find_total_pencapaian_row(self.penjualan_processed, "PENJUALAN")

            # Buat 3 baris total berdasarkan baris yang ditemukan
            print(f"\n🧮 Membuat 3 baris total...")

            # Baris 1: TOTAL KONSINYASI (copas dari data detail)
            row1 = None
            if kons_total_detail is not None:
                print(f"  ✓ Membuat TOTAL KONSINYASI dari baris yang ditemukan")
                row1 = self._create_total_row_from_detail("TOTAL KONSINYASI", kons_total_detail, ordered_skus, "konsinyasi")
            else:
                print(f"  ⚠ Tidak ada data TOTAL KONSINYASI untuk di-copas")

            # Baris 2: TOTAL PENJUALAN (copas dari data detail)  
            row2 = None
            if penj_total_detail is not None:
                print(f"  ✓ Membuat TOTAL PENJUALAN dari baris yang ditemukan")
                row2 = self._create_total_row_from_detail("TOTAL PENJUALAN", penj_total_detail, ordered_skus, "penjualan")
            else:
                print(f"  ⚠ Tidak ada data TOTAL PENJUALAN untuk di-copas")

            # Baris 3: TOTAL GABUNGAN (tambahan dari baris 1 + baris 2)
            row3 = None
            if row1 is not None or row2 is not None:
                print(f"  ✓ Membuat TOTAL GABUNGAN dari penjumlahan")
                row3 = self._create_gabungan_total_row(row1, row2, ordered_skus)
            else:
                print(f"  ❌ Tidak bisa membuat TOTAL GABUNGAN (data kosong)")
            

            # ================= 5. SORTING DATA GLOBAL BERDASARKAN PERINGKAT =================
            print(f"\n📊 SORTING GLOBAL BERDASARKAN PERINGKAT SKU (KONSINYASI+PENJUALAN)")

            # Buat list untuk menyimpan semua data dengan metadata untuk sorting
            all_data_with_meta = []

            # Kumpulkan semua data detail (bukan total) dengan metadata
            for row in all_rows:
                if len(row) > 1:  # Pastikan ada cukup data
                    nama = row[1] if len(row) > 1 else ""
                    jenis = row[0] if row else ""
                    
                    # Skip baris TOTAL untuk sorting (akan ditambahkan di akhir)
                    if "TOTAL" in str(nama).upper():
                        continue
                    
                    # Hitung jumlah SKU yang terjual
                    sku_terjual = 0
                    total_penjualan = 0
                    
                    # Index SKU dimulai setelah base headers (JENIS + NAMA + TOTAL PENJUALAN = 4 kolom)
                    sku_start_idx = 3
                    
                    # Hitung SKU yang terjual (nilai > 0)
                    for i in range(sku_start_idx, min(sku_start_idx + len(ordered_skus), len(row))):
                        value = row[i]
                        num_value = self._extract_number_from_value(value)
                        if num_value > 0:
                            sku_terjual += 1
                            total_penjualan += num_value
                    
                    # Dapatkan ranking dari kolom terakhir
                    ranking = row[-1] if len(row) > 0 else ""
                    
                    # Simpan dengan metadata untuk sorting
                    all_data_with_meta.append({
                        'row_data': row,
                        'jenis': jenis,
                        'nama': nama,
                        'sku_terjual': sku_terjual,
                        'total_penjualan': total_penjualan,
                        'ranking': ranking,
                        'has_ranking': ranking and str(ranking).strip() != "" and "Tidak ada pembelian" not in ranking
                    })

            # ================= LOGIKA SORTING BARU =================
            # Urutkan berdasarkan: 
            # 1. Ada ranking vs tidak ada ranking (TRUE lebih dulu)
            # 2. Jumlah SKU yang dibeli (descending)
            # 3. Total penjualan (descending)
            # 4. Jenis (KONSINYASI dulu)
            # 5. Nama (ascending)

            sorted_data = sorted(all_data_with_meta, key=lambda x: (
                # Prioritas 1: Punya ranking (True) vs tidak ada ranking (False)
                -1 if x['has_ranking'] else 0,
                # Prioritas 2: Jumlah SKU yang dibeli (descending)
                -x['sku_terjual'],
                # Prioritas 3: Total penjualan (descending)
                -x['total_penjualan'],
                # Prioritas 4: Jenis (KONSINYASI dulu)
                0 if x['jenis'] == 'KONSINYASI' else 1,
                # Prioritas 5: Nama (ascending)
                x['nama'].upper()
            ))

            # Tampilkan contoh hasil sorting
            print(f"\n📋 CONTOH HASIL SORTING GLOBAL:")
            for i, data in enumerate(sorted_data[:10]):
                print(f"  {i+1}. {data['jenis']} - {data['nama'][:30]}...")
                print(f"     • SKU terjual: {data['sku_terjual']} | Total: {data['total_penjualan']:,}")
                if data['ranking']:
                    ranking_preview = data['ranking'][:50] + "..." if len(data['ranking']) > 50 else data['ranking']
                    print(f"     • Peringkat: {ranking_preview}")

            # ================= SUSUN ULANG ALL_ROWS =================
            # 1. Tambahkan detail data yang sudah di-sort DENGAN RANKING GLOBAL
            all_rows_sorted = []
            for i, data in enumerate(sorted_data):  # ← PERUBAHAN: tambah enumerate
                row = data['row_data'].copy()
                
                # ====== OVERWRITE KOLOM "PERINGKAT SKU" DENGAN ANGKA RANKING ======
                # Kolom "PERINGKAT SKU" adalah kolom TERAKHIR
                index_peringkat_sku = len(row) - 1
                
                # Cek apakah ini baris TOTAL
                nama_pelanggan = row[1] if len(row) > 1 else ""
                is_total_row = "TOTAL" in str(nama_pelanggan).upper()
                
                if is_total_row:
                    # Baris TOTAL: ranking KOSONG
                    row[index_peringkat_sku] = ""
                else:
                    # Baris biasa: ranking angka (1, 2, 3, ...)
                    row[index_peringkat_sku] = i + 1  # ← ANGKA RANKING!
                
                all_rows_sorted.append(row)

            # 2. Tambahkan baris TOTAL di akhir (sesuai urutan di contoh)
            # TOTAL KONSINYASI
            if row1 is not None:
                print(f"  ✓ Menambahkan TOTAL KONSINYASI di akhir")
                if len(row1) < len(all_headers):
                    row1 = row1 + [""] * (len(all_headers) - len(row1))
                all_rows_sorted.append(row1)

            # TOTAL PENJUALAN
            if row2 is not None:
                print(f"  ✓ Menambahkan TOTAL PENJUALAN di akhir")
                if len(row2) < len(all_headers):
                    row2 = row2 + [""] * (len(all_headers) - len(row2))
                all_rows_sorted.append(row2)

            # TOTAL GABUNGAN
            if row3 is not None:
                print(f"  ✓ Menambahkan TOTAL GABUNGAN di akhir")
                if len(row3) < len(all_headers):
                    row3 = row3 + [""] * (len(all_headers) - len(row3))
                all_rows_sorted.append(row3)

            print(f"\n📊 HASIL AKHIR SORTING:")
            print(f"  - Detail (sorted): {len(sorted_data)} baris")
            print(f"  - Total (di akhir): {3 if (row1 or row2 or row3) else 0} baris")
            print(f"  - Total semua baris: {len(all_rows_sorted)}")

            # Ganti all_rows dengan yang sudah di-sort
            all_rows = all_rows_sorted
                    

            # ================= 6. TAMPILKAN DALAM 1 TABEL =================
            if all_rows and all_headers:
                print(f"\n✓ Menampilkan {len(all_rows)} baris dalam 1 tabel")
                print(f"  - Detail: {len([r for r in all_rows if r[0] == 'KONSINYASI'])} konsinyasi, "
                    f"{len([r for r in all_rows if r[0] == 'PENJUALAN'])} penjualan, "
                    f"{len([r for r in all_rows if 'TOTAL' in r[0]])} total")
                
                # Setup tabel
                self.table_gabungan.setRowCount(len(all_rows))
                self.table_gabungan.setColumnCount(len(all_headers))
                self.table_gabungan.setHorizontalHeaderLabels(all_headers)
                
                # Isi data dengan formatting
                for row_idx, row_data in enumerate(all_rows):
                    if len(row_data) != len(all_headers):
                        # Pad atau trim row data agar sesuai dengan header
                        if len(row_data) > len(all_headers):
                            row_data = row_data[:len(all_headers)]
                            print(f"⚠ Baris {row_idx} dipotong dari {len(row_data)} ke {len(all_headers)} kolom")
                        else:
                            missing = len(all_headers) - len(row_data)
                            row_data = row_data + [""] * missing
                            print(f"⚠ Baris {row_idx} ditambah {missing} kolom kosong")
                    
                    jenis = row_data[0] if row_data else ""  # Kolom pertama adalah JENIS
                    
                    for col_idx, cell_value in enumerate(row_data):
                        col_name = all_headers[col_idx] if col_idx < len(all_headers) else ""
                        
                        # Handle NaN values
                        if pd.isna(cell_value):
                            item = QTableWidgetItem("")
                        else:
                            item = QTableWidgetItem(str(cell_value))
                        
                        # Format berdasarkan jenis data
                        if isinstance(cell_value, (int, float)) and not pd.isna(cell_value):
                            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                            if cell_value == 0:
                                item.setText("0")
                            elif isinstance(cell_value, int) or (isinstance(cell_value, float) and cell_value.is_integer()):
                                item.setText(f"{int(cell_value):,}")
                            else:
                                item.setText(f"{cell_value:,.2f}")
                        else:
                            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                        
                        # Beri warna berdasarkan JENIS
                        if jenis == "KONSINYASI":
                            if row_idx % 2 == 0:
                                item.setBackground(QColor(232, 244, 248))  # Biru muda
                            else:
                                item.setBackground(QColor(242, 250, 252))  # Biru lebih muda
                            item.setFont(QFont("Arial", 9))
                        
                        elif jenis == "PENJUALAN":
                            if row_idx % 2 == 0:
                                item.setBackground(QColor(248, 232, 244))  # Pink muda
                            else:
                                item.setBackground(QColor(252, 242, 250))  # Pink lebih muda
                            item.setFont(QFont("Arial", 9))
                        
                        elif "TOTAL" in str(jenis):
                            if "KONSINYASI" in str(jenis):
                                item.setBackground(QColor(173, 216, 230))  # Light Blue
                                item.setFont(QFont("Arial", 9, QFont.Bold))
                            elif "PENJUALAN" in str(jenis):
                                item.setBackground(QColor(255, 182, 193))  # Light Pink
                                item.setFont(QFont("Arial", 9, QFont.Bold))
                            elif "GABUNGAN" in str(jenis):
                                item.setBackground(QColor(255, 255, 200))  # Light Yellow
                                item.setFont(QFont("Arial", 10, QFont.Bold))
                            
                            # Highlight khusus untuk kolom statistik SKU
                            if col_name and self._is_sku_stat_column(col_name):
                                if "KONSINYASI" in str(jenis):
                                    item.setBackground(QColor(200, 230, 255))  # Blue lighter
                                elif "PENJUALAN" in str(jenis):
                                    item.setBackground(QColor(255, 200, 210))  # Pink lighter
                                elif "GABUNGAN" in str(jenis):
                                    item.setBackground(QColor(255, 255, 220))  # Yellow lighter
                            
                            # Highlight kolom NAMA PELANGGAN untuk total
                            if col_name and ('NAMA' in col_name.upper() or 'PELANGGAN' in col_name.upper()):
                                if "KONSINYASI" in str(jenis):
                                    item.setForeground(QColor(41, 128, 185))  # Blue
                                elif "PENJUALAN" in str(jenis):
                                    item.setForeground(QColor(142, 68, 173))  # Purple
                                elif "GABUNGAN" in str(jenis):
                                    item.setForeground(QColor(39, 174, 96))  # Green
                        
                        # Highlight kolom SKU dengan bold
                        if col_name and (self._is_pure_sku_column(col_name) or 
                                    str(col_name).upper() in [s.upper() for s in self.all_skus]):
                            font = item.font()
                            font.setBold(True)
                            item.setFont(font)
                            item.setForeground(QColor(155, 89, 182))  # Ungu untuk SKU
                        
                        # Format khusus untuk kolom PERINGKAT SKU
                        if col_name and 'PERINGKAT SKU' in str(col_name).upper():
                            # Format khusus untuk kolom peringkat
                            item.setFont(QFont("Arial", 9))
                            
                            # Warna berdasarkan ranking
                            cell_text = str(cell_value).upper()
                            if '1.' in cell_text:
                                item.setForeground(QColor(230, 126, 34))  # Orange untuk peringkat 1
                                font = item.font()
                                font.setBold(True)
                                item.setFont(font)
                            elif '2.' in cell_text:
                                item.setForeground(QColor(52, 152, 219))  # Blue untuk peringkat 2
                            elif '3.' in cell_text:
                                item.setForeground(QColor(155, 89, 182))  # Purple untuk peringkat 3
                            
                            # Untuk baris tanpa ranking (termasuk TOTAL)
                            if not cell_text or cell_text == "":
                                item.setForeground(QColor(149, 165, 166))  # Gray untuk kosong
                        
                        self.table_gabungan.setItem(row_idx, col_idx, item)
                
                # Auto resize columns
                self.table_gabungan.resizeColumnsToContents()
                
                # Set width khusus untuk kolom statistik SKU
                for col in range(self.table_gabungan.columnCount()):
                    if col < len(all_headers):
                        col_name = str(all_headers[col]).upper()
                        
                        if self._is_sku_stat_column(col_name):
                            if 'SKU YANG' in col_name:
                                self.table_gabungan.setColumnWidth(col, 200)
                            elif 'JML SKU' in col_name:
                                self.table_gabungan.setColumnWidth(col, 100)
                            elif col_name == '%':
                                self.table_gabungan.setColumnWidth(col, 80)
                            elif 'PERINGKAT' in col_name:
                                self.table_gabungan.setColumnWidth(col, 250)  # Lebar lebih untuk ranking
                        else:
                            width = self.table_gabungan.columnWidth(col)
                            if width < 80:
                                self.table_gabungan.setColumnWidth(col, 80)
                
                # Update stats label
                kons_count = sum(1 for row in all_rows if row[0] == "KONSINYASI")
                penj_count = sum(1 for row in all_rows if row[0] == "PENJUALAN")
                total_count = sum(1 for row in all_rows if "TOTAL" in str(row[0]))
                
                self.stats_label.setText(
                    f"📊 GABUNGAN: {kons_count} konsinyasi + {penj_count} penjualan + {total_count} total | "
                    f"Total baris: {len(all_rows)} | "
                    f"SKU: {len(ordered_skus)} (default: {len(self.default_skus)}, dynamic: {len(self.all_dynamic_skus)})"
                )
                
                # Tampilkan konfirmasi di status bar
                self.status_bar.showMessage(
                    f"✓ Tampilan gabungan diperbarui: {kons_count} konsinyasi, {penj_count} penjualan, {total_count} total"
                )
                
            else:
                print(f"⚠ Tidak ada data untuk ditampilkan")
                self.table_gabungan.clear()
                self.table_gabungan.setRowCount(0)
                self.table_gabungan.setColumnCount(0)
                self.stats_label.setText("📊 Tidak ada data untuk ditampilkan")
            
            # Aktifkan tombol export jika ada data
            has_kons = self.konsinyasi_processed is not None and not self.konsinyasi_processed.empty
            has_penj = self.penjualan_processed is not None and not self.penjualan_processed.empty
            self.export_btn.setEnabled(has_kons or has_penj)
            
            print(f"✓ Selesai: {len(all_rows)} baris ditampilkan")
            
        except Exception as e:
            print(f"❌ Error in update_gabungan_view: {e}")
            import traceback
            traceback.print_exc()
            self.status_bar.showMessage(f"❌ Error: {str(e)}")
            
                # Fungsi untuk mencari baris TOTAL PENCAPAIAN
            def find_total_pencapaian_row(df, jenis):
                if df is None or df.empty:
                    return None
                
                # Cari di semua kolom yang mengandung 'NAMA' atau 'PELANGGAN'
                nama_cols = [col for col in df.columns if any(keyword in str(col).upper() for keyword in ['NAMA', 'PELANGGAN', 'CUSTOMER'])]
                
                for col in nama_cols:
                    mask = df[col].astype(str).str.contains('TOTAL PENCAPAIAN', case=False, na=False)
                    if mask.any():
                        print(f"  ✓ TOTAL PENCAPAIAN ditemukan di {jenis} (kolom: {col})")
                        return df[mask].iloc[0].copy()
                
                # Jika tidak ditemukan, coba cari baris dengan kata "TOTAL" di kolom pertama
                first_col = df.columns[0]
                mask = df[first_col].astype(str).str.contains('TOTAL', case=False, na=False)
                if mask.any():
                    print(f"  ⚠ TOTAL (tanpa PENCAPAIAN) ditemukan di {jenis}")
                    return df[mask].iloc[0].copy()
                
                print(f"  ❌ Tidak ditemukan baris TOTAL di {jenis}")
                return None

            # Cari baris TOTAL PENCAPAIAN dari masing-masing data detail
            kons_total_detail = None
            penj_total_detail = None

            if self.konsinyasi_processed is not None:
                kons_total_detail = find_total_pencapaian_row(self.konsinyasi_processed, "KONSINYASI")

            if self.penjualan_processed is not None:
                penj_total_detail = find_total_pencapaian_row(self.penjualan_processed, "PENJUALAN")

            # Buat 3 baris total berdasarkan baris yang ditemukan
            print(f"\n🧮 Membuat 3 baris total...")

            # Baris 1: TOTAL KONSINYASI (copas dari data detail)
            row1 = None
            if kons_total_detail is not None:
                print(f"  ✓ Membuat TOTAL KONSINYASI dari baris yang ditemukan")
                row1 = self._create_total_row_from_detail("TOTAL KONSINYASI", kons_total_detail, ordered_skus, "konsinyasi")
            else:
                print(f"  ⚠ Tidak ada data TOTAL KONSINYASI untuk di-copas")

            # Baris 2: TOTAL PENJUALAN (copas dari data detail)  
            row2 = None
            if penj_total_detail is not None:
                print(f"  ✓ Membuat TOTAL PENJUALAN dari baris yang ditemukan")
                row2 = self._create_total_row_from_detail("TOTAL PENJUALAN", penj_total_detail, ordered_skus, "penjualan")
            else:
                print(f"  ⚠ Tidak ada data TOTAL PENJUALAN untuk di-copas")

            # Baris 3: TOTAL GABUNGAN (tambahan dari baris 1 + baris 2)
            row3 = None
            if row1 is not None or row2 is not None:
                print(f"  ✓ Membuat TOTAL GABUNGAN dari penjumlahan")
                row3 = self._create_gabungan_total_row(row1, row2, ordered_skus)
            else:
                print(f"  ❌ Tidak bisa membuat TOTAL GABUNGAN (data kosong)")

            # ================= 4. TAMBAHKAN 3 BARIS TOTAL KE ALL_ROWS =================
            print(f"\n📊 Menambahkan 3 baris total ke tabel...")

            if row1 is not None:
                print(f"  ✓ Menambahkan TOTAL KONSINYASI")
                all_rows.append(row1)

            if row2 is not None:
                print(f"  ✓ Menambahkan TOTAL PENJUALAN")
                all_rows.append(row2)

            if row3 is not None:
                print(f"  ✓ Menambahkan TOTAL GABUNGAN")
                all_rows.append(row3)
                        
            # ================= 5. TAMPILKAN DALAM 1 TABEL =================
            if all_rows and all_headers:
                print(f"\n✓ Menampilkan {len(all_rows)} baris dalam 1 tabel")
                print(f"  - Detail: {len([r for r in all_rows if r[0] == 'KONSINYASI'])} konsinyasi, "
                    f"{len([r for r in all_rows if r[0] == 'PENJUALAN'])} penjualan, "
                    f"{len([r for r in all_rows if 'TOTAL' in r[0]])} total")
                
                # Setup tabel
                self.table_gabungan.setRowCount(len(all_rows))
                self.table_gabungan.setColumnCount(len(all_headers))
                self.table_gabungan.setHorizontalHeaderLabels(all_headers)
                
                # Isi data dengan formatting
                for row_idx, row_data in enumerate(all_rows):
                    if len(row_data) != len(all_headers):
                        # Pad atau trim row data agar sesuai dengan header
                        if len(row_data) > len(all_headers):
                            row_data = row_data[:len(all_headers)]
                            print(f"⚠ Baris {row_idx} dipotong dari {len(row_data)} ke {len(all_headers)} kolom")
                        else:
                            missing = len(all_headers) - len(row_data)
                            row_data = row_data + [""] * missing
                            print(f"⚠ Baris {row_idx} ditambah {missing} kolom kosong")
                    
                    jenis = row_data[0] if row_data else ""  # Kolom pertama adalah JENIS
                    
                    for col_idx, cell_value in enumerate(row_data):
                        col_name = all_headers[col_idx] if col_idx < len(all_headers) else ""
                        
                        # Handle NaN values
                        if pd.isna(cell_value):
                            item = QTableWidgetItem("")
                        else:
                            item = QTableWidgetItem(str(cell_value))
                        
                        # Format berdasarkan jenis data
                        if isinstance(cell_value, (int, float)) and not pd.isna(cell_value):
                            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                            if cell_value == 0:
                                item.setText("0")
                            elif isinstance(cell_value, int) or (isinstance(cell_value, float) and cell_value.is_integer()):
                                item.setText(f"{int(cell_value):,}")
                            else:
                                item.setText(f"{cell_value:,.2f}")
                        else:
                            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                        
                        # Beri warna berdasarkan JENIS
                        if jenis == "KONSINYASI":
                            if row_idx % 2 == 0:
                                item.setBackground(QColor(232, 244, 248))  # Biru muda
                            else:
                                item.setBackground(QColor(242, 250, 252))  # Biru lebih muda
                            item.setFont(QFont("Arial", 9))
                        
                        elif jenis == "PENJUALAN":
                            if row_idx % 2 == 0:
                                item.setBackground(QColor(248, 232, 244))  # Pink muda
                            else:
                                item.setBackground(QColor(252, 242, 250))  # Pink lebih muda
                            item.setFont(QFont("Arial", 9))
                        
                        elif "TOTAL" in str(jenis):
                            if "KONSINYASI" in str(jenis):
                                item.setBackground(QColor(173, 216, 230))  # Light Blue
                                item.setFont(QFont("Arial", 9, QFont.Bold))
                            elif "PENJUALAN" in str(jenis):
                                item.setBackground(QColor(255, 182, 193))  # Light Pink
                                item.setFont(QFont("Arial", 9, QFont.Bold))
                            elif "GABUNGAN" in str(jenis):
                                item.setBackground(QColor(255, 255, 200))  # Light Yellow
                                item.setFont(QFont("Arial", 10, QFont.Bold))
                            
                            # Highlight khusus untuk kolom statistik SKU
                            if col_name and self._is_sku_stat_column(col_name):
                                if "KONSINYASI" in str(jenis):
                                    item.setBackground(QColor(200, 230, 255))  # Blue lighter
                                elif "PENJUALAN" in str(jenis):
                                    item.setBackground(QColor(255, 200, 210))  # Pink lighter
                                elif "GABUNGAN" in str(jenis):
                                    item.setBackground(QColor(255, 255, 220))  # Yellow lighter
                            
                            # Highlight kolom NAMA PELANGGAN untuk total
                            if col_name and ('NAMA' in col_name.upper() or 'PELANGGAN' in col_name.upper()):
                                if "KONSINYASI" in str(jenis):
                                    item.setForeground(QColor(41, 128, 185))  # Blue
                                elif "PENJUALAN" in str(jenis):
                                    item.setForeground(QColor(142, 68, 173))  # Purple
                                elif "GABUNGAN" in str(jenis):
                                    item.setForeground(QColor(39, 174, 96))  # Green
                        
                        # Highlight kolom SKU dengan bold
                        if col_name and (self._is_pure_sku_column(col_name) or 
                                    str(col_name).upper() in [s.upper() for s in self.all_skus]):
                            font = item.font()
                            font.setBold(True)
                            item.setFont(font)
                            item.setForeground(QColor(155, 89, 182))  # Ungu untuk SKU
                        
                        self.table_gabungan.setItem(row_idx, col_idx, item)
                
                # Auto resize columns
                self.table_gabungan.resizeColumnsToContents()
                
                # Set width khusus untuk kolom statistik SKU
                for col in range(self.table_gabungan.columnCount()):
                    if col < len(all_headers):
                        col_name = str(all_headers[col]).upper()
                        
                        if self._is_sku_stat_column(col_name):
                            if 'SKU YANG' in col_name:
                                self.table_gabungan.setColumnWidth(col, 200)
                            elif 'JML SKU' in col_name:
                                self.table_gabungan.setColumnWidth(col, 100)
                            elif col_name == '%':
                                self.table_gabungan.setColumnWidth(col, 80)
                            elif 'PERINGKAT' in col_name:
                                self.table_gabungan.setColumnWidth(col, 120)
                        else:
                            width = self.table_gabungan.columnWidth(col)
                            if width < 80:
                                self.table_gabungan.setColumnWidth(col, 80)
                
                # Update stats label
                kons_count = sum(1 for row in all_rows if row[0] == "KONSINYASI")
                penj_count = sum(1 for row in all_rows if row[0] == "PENJUALAN")
                total_count = sum(1 for row in all_rows if "TOTAL" in str(row[0]))
                
                self.stats_label.setText(
                    f"📊 GABUNGAN: {kons_count} konsinyasi + {penj_count} penjualan + {total_count} total | "
                    f"Total baris: {len(all_rows)} | "
                    f"SKU: {len(ordered_skus)} (default: {len(self.default_skus)}, dynamic: {len(self.all_dynamic_skus)})"
                )
                
                # Tampilkan konfirmasi di status bar
                self.status_bar.showMessage(
                    f"✓ Tampilan gabungan diperbarui: {kons_count} konsinyasi, {penj_count} penjualan, {total_count} total"
                )
                
            else:
                print(f"⚠ Tidak ada data untuk ditampilkan")
                self.table_gabungan.clear()
                self.table_gabungan.setRowCount(0)
                self.table_gabungan.setColumnCount(0)
                self.stats_label.setText("📊 Tidak ada data untuk ditampilkan")
            
            # Aktifkan tombol export jika ada data
            has_kons = self.konsinyasi_processed is not None and not self.konsinyasi_processed.empty
            has_penj = self.penjualan_processed is not None and not self.penjualan_processed.empty
            self.export_btn.setEnabled(has_kons or has_penj)
            
            print(f"✓ Selesai: {len(all_rows)} baris ditampilkan")
            
        except Exception as e:
            print(f"❌ Error in update_gabungan_view: {e}")
            import traceback
            traceback.print_exc()
            self.status_bar.showMessage(f"❌ Error: {str(e)}")

    def _create_total_row_from_detail(self, jenis, detail_row, ordered_skus, source_type):
        """Buat baris total dari baris detail TOTAL PENCAPAIAN"""
        try:
            print(f"  🔧 Membuat baris {jenis} dari detail...")
            
            # Baris baru dimulai dengan JENIS
            new_row = [jenis]
            
            # Dapatkan DataFrame yang sesuai untuk perhitungan statistik
            if source_type == "konsinyasi":
                df_for_stats = self.konsinyasi_processed
            else:
                df_for_stats = self.penjualan_processed
            
            # 1. Tambahkan NAMA PELANGGAN (deskripsi)
            new_row.append("TOTAL PENCAPAIAN")
            
            # **HAPUS KOLOM SISTEM SEPENUHNYA - tidak diperlukan**
            # new_row.append(system_name)  # <-- INI HARUS DIHAPUS TOTAL
            
            # 2. Hitung TOTAL PENJUALAN dari detail_row
            total_penjualan = 0
            # Cari kolom yang berisi TOTAL PENJUALAN
            for col_name, value in detail_row.items():
                col_str = str(col_name).upper()
                if 'TOTAL' in col_str and 'PENJUALAN' in col_str:
                    total_penjualan = self._extract_number_from_value(value)
                    break
            new_row.append(total_penjualan)
            
            # 3. Tambahkan nilai untuk setiap SKU dari detail_row
            sku_masuk_list = []
            sku_tidak_masuk_list = []
            
            for sku in ordered_skus:
                # Cari nilai di detail_row
                sku_value = 0
                # Cari dengan case-insensitive
                for col_name, value in detail_row.items():
                    if str(col_name).upper().strip() == sku.upper():
                        sku_value = self._extract_number_from_value(value)
                        break
                
                new_row.append(sku_value)
                
                # Catat untuk statistik awal
                if sku_value > 0:
                    sku_masuk_list.append(sku)
                else:
                    sku_tidak_masuk_list.append(sku)
            
            # 4. Hitung statistik SKU dari data lengkap
            if df_for_stats is not None and not df_for_stats.empty:
                # Hitung ulang statistik dari data sebenarnya
                sku_masuk_list = []  # Reset dan hitung ulang
                sku_tidak_masuk_list = []  # Reset dan hitung ulang
                
                for sku in ordered_skus:
                    if sku in df_for_stats.columns:
                        try:
                            # Konversi ke numeric
                            df_sku_numeric = pd.to_numeric(df_for_stats[sku], errors='coerce')
                            # Hitung berapa pelanggan yang membeli SKU ini (nilai > 0)
                            pelanggan_dengan_sku = (df_sku_numeric > 0).sum()
                            
                            if pelanggan_dengan_sku > 0:
                                sku_masuk_list.append(sku)
                            else:
                                sku_tidak_masuk_list.append(sku)
                                
                        except Exception as e:
                            sku_tidak_masuk_list.append(sku)
                    else:
                        sku_tidak_masuk_list.append(sku)
                
                print(f"    ✓ Data lengkap untuk statistik: {len(df_for_stats)} pelanggan")
                
            else:
                print(f"    ⚠ Tidak ada data lengkap, gunakan perhitungan dari detail")
            
            total_sku = len(ordered_skus)
            sku_masuk_count = len(sku_masuk_list)
            persen_masuk = (sku_masuk_count / total_sku * 100) if total_sku > 0 else 0
            persen_tidak_masuk = 100 - persen_masuk
            
            print(f"    ✓ {jenis}: {sku_masuk_count}/{total_sku} SKU terjual ({persen_masuk:.1f}%)")
            
            # 5. Format string untuk kolom statistik SKU
            sku_masuk_str = ", ".join(sku_masuk_list) if sku_masuk_list else "-"
            sku_tidak_str = ", ".join(sku_tidak_masuk_list) if sku_tidak_masuk_list else "-"
            
            # 6. Tambahkan kolom statistik SKU
            new_row.extend([
                sku_masuk_str,          # SKU YANG SUDAH MASUK
                sku_tidak_str,          # SKU YANG BELUM MASUK
                f"{sku_masuk_count}/{total_sku}",  # JUMLAH SKU MASUK
                f"{persen_masuk:.1f}%",  # %
                f"{len(sku_tidak_masuk_list)}/{total_sku}",  # JUMLAH SKU TIDAK MASUK
                f"{persen_tidak_masuk:.1f}%",  # %
                ""  # PERINGKAT SKU (KOSONG untuk total - sesuai permintaan)
            ])
            
            return new_row
            
        except Exception as e:
            print(f"❌ Error creating total row from detail: {e}")
            import traceback
            traceback.print_exc()
            return None
            
    def _create_gabungan_total_row(self, row1, row2, ordered_skus):
        """Buat baris TOTAL GABUNGAN dari penjumlahan row1 dan row2"""
        try:
            print(f"  🔧 Membuat baris TOTAL GABUNGAN...")
            
            # Baris baru dimulai dengan JENIS
            new_row = ["TOTAL GABUNGAN"]
            
            # 1. Tambahkan NAMA PELANGGAN (deskripsi)
            new_row.append("TOTAL PENCAPAIAN SEMUA")
            
            # **PERBAIKAN: HITUNG TOTAL PENJUALAN GABUNGAN DENGAN BENAR**
            total_penjualan_gabungan = 0
            
            # Ambil nilai dari row1 (KONSINYASI)
            if row1 and len(row1) > 2:  # row1[2] adalah TOTAL PENJUALAN (setelah JENIS dan NAMA)
                total_kons = self._extract_number_from_value(row1[2])
                total_penjualan_gabungan += total_kons
                print(f"    - TOTAL KONSINYASI: {total_kons}")
            
            # Ambil nilai dari row2 (PENJUALAN)
            if row2 and len(row2) > 2:  # row2[2] adalah TOTAL PENJUALAN
                total_penj = self._extract_number_from_value(row2[2])
                total_penjualan_gabungan += total_penj
                print(f"    - TOTAL PENJUALAN: {total_penj}")
            
            new_row.append(total_penjualan_gabungan)
            print(f"    ✓ TOTAL GABUNGAN: {total_penjualan_gabungan}")
            
            # 2. Hitung nilai untuk setiap SKU (jumlah dari kedua sistem)
            sku_masuk_list = []
            sku_tidak_masuk_list = []
            
            # **PERBAIKAN: Index mulai untuk kolom SKU**
            # Setelah 3 kolom pertama: JENIS (0), NAMA PELANGGAN (1), TOTAL PENJUALAN (2)
            start_idx = 3
            
            for i, sku in enumerate(ordered_skus):
                idx = start_idx + i
                
                # Ambil nilai dari kedua baris
                kons_val = 0
                penj_val = 0
                
                if row1 and idx < len(row1):
                    kons_val = self._extract_number_from_value(row1[idx])
                
                if row2 and idx < len(row2):
                    penj_val = self._extract_number_from_value(row2[idx])
                
                # Jumlahkan
                total_sku_value = kons_val + penj_val
                new_row.append(total_sku_value)
                
                # Catat untuk statistik
                if total_sku_value > 0:
                    sku_masuk_list.append(sku)
                else:
                    sku_tidak_masuk_list.append(sku)
            
            # 3. Hitung statistik SKU untuk GABUNGAN
            # Cek apakah SKU terjual di salah satu sistem
            sku_masuk_list = []  # Reset
            sku_tidak_masuk_list = []  # Reset
            
            # Data untuk pengecekan
            kons_terjual_dict = {}
            penj_terjual_dict = {}
            
            # Hitung untuk konsinyasi
            if self.konsinyasi_processed is not None:
                for sku in ordered_skus:
                    if sku in self.konsinyasi_processed.columns:
                        try:
                            kons_series = pd.to_numeric(self.konsinyasi_processed[sku], errors='coerce')
                            kons_terjual_dict[sku] = (kons_series > 0).any()
                        except:
                            kons_terjual_dict[sku] = False
            
            # Hitung untuk penjualan
            if self.penjualan_processed is not None:
                for sku in ordered_skus:
                    if sku in self.penjualan_processed.columns:
                        try:
                            penj_series = pd.to_numeric(self.penjualan_processed[sku], errors='coerce')
                            penj_terjual_dict[sku] = (penj_series > 0).any()
                        except:
                            penj_terjual_dict[sku] = False
            
            # Tentukan SKU yang masuk/tidak masuk
            for sku in ordered_skus:
                kons_terjual = kons_terjual_dict.get(sku, False)
                penj_terjual = penj_terjual_dict.get(sku, False)
                
                if kons_terjual or penj_terjual:
                    sku_masuk_list.append(sku)
                else:
                    sku_tidak_masuk_list.append(sku)
            
            total_sku = len(ordered_skus)
            sku_masuk_count = len(sku_masuk_list)
            persen_masuk = (sku_masuk_count / total_sku * 100) if total_sku > 0 else 0
            persen_tidak_masuk = 100 - persen_masuk
            
            print(f"    ✓ GABUNGAN: {sku_masuk_count}/{total_sku} SKU terjual ({persen_masuk:.1f}%)")
            print(f"      - SKU terjual di KONSINYASI atau PENJUALAN")
            
            # 4. Format string untuk kolom statistik SKU
            sku_masuk_str = ", ".join(sku_masuk_list) if sku_masuk_list else "-"
            sku_tidak_str = ", ".join(sku_tidak_masuk_list) if sku_tidak_masuk_list else "-"
            
            # 5. Tambahkan kolom statistik SKU
            new_row.extend([
                sku_masuk_str,          # SKU YANG SUDAH MASUK
                sku_tidak_str,          # SKU YANG BELUM MASUK
                f"{sku_masuk_count}/{total_sku}",  # JUMLAH SKU MASUK
                f"{persen_masuk:.1f}%",  # %
                f"{len(sku_tidak_masuk_list)}/{total_sku}",  # JUMLAH SKU TIDAK MASUK
                f"{persen_tidak_masuk:.1f}%",  # %
                ""  # PERINGKAT SKU (kosong untuk total)
            ])
            
            return new_row
            
        except Exception as e:
            print(f"❌ Error creating gabungan total row: {e}")
            import traceback
            traceback.print_exc()
            return None

    def detect_cabang_from_filename(self, file_name):
        """Deteksi cabang dari nama file - EXACT MATCH"""
        file_upper = file_name.upper()
        
        # Pattern: cari "GJK" sebagai substring terpisah
        import re
        
        # Cari pattern kode 2-3 huruf setelah "LPD" atau di nama file
        patterns = [
            r'LPD\s*[-_]\s*([A-Z]{2,3})',  # LPD - GJK
            r'([A-Z]{2,3})\.xlsx',          # GJK.xlsx
            r'([A-Z]{2,3})\s',              # GJK (diikuti spasi)
            r'\s([A-Z]{2,3})\s',            # spasi GJK spasi
        ]
        
        for pattern in patterns:
            match = re.search(pattern, file_upper)
            if match:
                kode = match.group(1)
                nama = CABANG_KODE_MAP.get(kode)
                if nama:
                    print(f"✓ Cabang terdeteksi: {kode} -> {nama}")
                    return nama, kode
        
        # Fallback: cari substring (tapi urut dari panjang ke pendek)
        sorted_keys = sorted(CABANG_KODE_MAP.keys(), key=len, reverse=True)
        for kode in sorted_keys:
            if kode in file_upper:
                nama = CABANG_KODE_MAP[kode]
                print(f"✓ Cabang terdeteksi (substring): {kode} -> {nama}")
                return nama, kode
        
        print(f"⚠ Cabang tidak terdeteksi, default ke JAKARTA")
        return "JAKARTA", "GJK"


    def setup_menu(self):
        """Setup menu bar"""
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("📁 File")
        
        # Tambah menu untuk load file masing-masing
        load_kons_action = QAction("📂 Buka File Konsinyasi...", self)
        load_kons_action.triggered.connect(lambda: self.browse_file("konsinyasi"))
        file_menu.addAction(load_kons_action)
        
        load_penj_action = QAction("📂 Buka File Penjualan...", self)
        load_penj_action.triggered.connect(lambda: self.browse_file("penjualan"))
        file_menu.addAction(load_penj_action)
        
        file_menu.addSeparator()
        
        export_action = QAction("💾 Export Laporan...", self)
        export_action.setShortcut("Ctrl+E")
        export_action.triggered.connect(self.export_data)
        file_menu.addAction(export_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("🚪 Keluar", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # View menu dengan shortcut
        view_menu = menubar.addMenu("👁️ View")
        
        view_konsinyasi_action = QAction("📋 View Konsinyasi", self)
        view_konsinyasi_action.setShortcut("Ctrl+1")
        view_konsinyasi_action.triggered.connect(lambda: self.switch_view_mode(0))
        view_menu.addAction(view_konsinyasi_action)
        
        view_penjualan_action = QAction("💰 View Penjualan", self)
        view_penjualan_action.setShortcut("Ctrl+2")
        view_penjualan_action.triggered.connect(lambda: self.switch_view_mode(1))
        view_menu.addAction(view_penjualan_action)
        
        view_gabungan_action = QAction("👁️ View Gabungan", self)
        view_gabungan_action.setShortcut("Ctrl+3")
        view_gabungan_action.triggered.connect(lambda: self.switch_view_mode(2))
        view_menu.addAction(view_gabungan_action)
        
        view_menu.addSeparator()
        
        auto_resize_action = QAction("🔍 Auto Resize Columns", self)
        auto_resize_action.setShortcut("Ctrl+R")
        auto_resize_action.triggered.connect(self.auto_resize_all_tables)
        view_menu.addAction(auto_resize_action)
        
        # Tools menu
        tools_menu = menubar.addMenu("⚙️ Tools")
        
        clear_action = QAction("🗑️ Clear All Data", self)
        clear_action.setShortcut("Ctrl+Shift+C")
        clear_action.triggered.connect(self.clear_all_data)
        tools_menu.addAction(clear_action)
        
        refresh_action = QAction("🔄 Refresh", self)
        refresh_action.setShortcut("F5")
        refresh_action.triggered.connect(self.refresh_data)
        tools_menu.addAction(refresh_action)
        
        # Help menu
        help_menu = menubar.addMenu("❓ Help")
        
        about_action = QAction("ℹ️ Tentang", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
            
    def setup_shortcuts(self):
        """Setup keyboard shortcuts"""
        refresh_shortcut = QShortcut("F5", self)
        refresh_shortcut.activated.connect(self.refresh_data)
        
        escape_shortcut = QShortcut("Escape", self)
        escape_shortcut.activated.connect(self.clear_table_selection)


    def create_file_input_panel(self, title, color):
        """Membuat panel input file untuk jenis laporan tertentu"""
        panel = QFrame()
        panel.setFrameShape(QFrame.StyledPanel)
        panel.setStyleSheet(f"""
            QFrame {{
                background:{color};
                border:1px solid #ddd;
                border-radius:5px;
                padding:10px;
            }}
        """)
        
        layout = QHBoxLayout(panel)
        layout.addWidget(QLabel(f"📁 {title}:"))
        
        file_label = QLabel("Belum ada file dipilih")
        file_label.setObjectName("file_label")
        file_label.setStyleSheet("color:#666; font-style:italic;")
        layout.addWidget(file_label, 1)
        
        browse_btn = QPushButton("📂 Pilih File")
        browse_btn.setObjectName("browse_btn")
        layout.addWidget(browse_btn)
        
        process_btn = QPushButton("⚙️ Proses")
        process_btn.setObjectName("process_btn")
        process_btn.setEnabled(False)
        layout.addWidget(process_btn)
        
        return panel
            
    def get_view_button_style(self, is_active):
        """Style untuk tombol view"""
        if is_active:
            return """
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    font-weight: bold;
                    border: 2px solid #2980b9;
                    border-radius: 5px;
                    padding: 8px 15px;
                }
                QPushButton:hover {
                    background-color: #2980b9;
                }
            """
        else:
            return """
                QPushButton {
                    background-color: #ecf0f1;
                    color: #7f8c8d;
                    font-weight: bold;
                    border: 2px solid #bdc3c7;
                    border-radius: 5px;
                    padding: 8px 15px;
                }
                QPushButton:hover {
                    background-color: #d5dbdb;
                }
            """

    def switch_view_mode(self, mode):
        """Ganti mode view (0=konsinyasi, 1=penjualan, 2=gabungan)"""
        self.current_view_mode = mode
        
        # Update tombol
        self.view_konsinyasi_btn.setChecked(mode == 0)
        self.view_penjualan_btn.setChecked(mode == 1)
        self.view_gabungan_btn.setChecked(mode == 2)
        
        # Update style tombol
        self.view_konsinyasi_btn.setStyleSheet(self.get_view_button_style(mode == 0))
        self.view_penjualan_btn.setStyleSheet(self.get_view_button_style(mode == 1))
        self.view_gabungan_btn.setStyleSheet(self.get_view_button_style(mode == 2))
        
        # Update mode label
        mode_labels = ["KONSINYASI", "PENJUALAN", "GABUNGAN"]
        self.mode_info_label.setText(f"Mode: {mode_labels[mode]}")
        self.mode_info_label.setStyleSheet("""
            color: white;
            background-color: #3498db;
            padding: 5px 10px;
            border-radius: 10px;
        """)
        
        # Switch tabel
        self.table_area.setCurrentIndex(mode)
        
        # Update data di view
        self.update_current_view()

    def browse_file(self, jenis):
        """Browse file berdasarkan jenis (konsinyasi/penjualan)"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            f"Pilih File {jenis.upper()}",
            "",
            "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        
        if file_path:
            self.load_file(jenis, file_path)

    def load_file(self, jenis, file_path):
        """Load file ke storage sesuai jenis - HANYA LOAD, TIDAK PROSES"""
        try:
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(20)
            
            print(f"\n📂 LOADING FILE {jenis}: {os.path.basename(file_path)}")
            
            # Baca file TANPA processing
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, encoding='utf-8')
            else:
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names
                df = excel_file.parse(sheet_names[0]) if sheet_names else pd.DataFrame()
            
            print(f"✓ File berhasil dibaca: {len(df)} baris, {len(df.columns)} kolom")
            
            self.progress_bar.setValue(60)
            
            # Simpan ke storage yang sesuai
            if jenis == "konsinyasi":
                self.konsinyasi_df = df
                self.konsinyasi_file = file_path
                self.kons_file_label.setText(f"✓ {os.path.basename(file_path)} ({len(df)} baris)")
                self.kons_file_label.setStyleSheet("color:#27ae60;")
                self.kons_process_btn.setEnabled(True)
                print(f"  - Disimpan sebagai data KONSINYASI")
            else:  # penjualan
                self.penjualan_df = df
                self.penjualan_file = file_path
                self.penj_file_label.setText(f"✓ {os.path.basename(file_path)} ({len(df)} baris)")
                self.penj_file_label.setStyleSheet("color:#27ae60;")
                self.penj_process_btn.setEnabled(True)
                print(f"  - Disimpan sebagai data PENJUALAN")
            
            self.progress_bar.setValue(80)
            
            # Tampilkan preview data di stats
            preview_cols = list(df.columns)[:5]
            self.stats_label.setText(
                f"📊 {jenis.upper()}: {len(df)} baris, {len(df.columns)} kolom | Kolom: {', '.join(preview_cols)}..."
            )
            
            # Update view TANPA processing
            self.update_current_view()
            
            self.progress_bar.setValue(100)
            QTimer.singleShot(500, lambda: self.progress_bar.setVisible(False))
            
            self.status_bar.showMessage(f"✓ File {jenis} berhasil dimuat: {os.path.basename(file_path)}")
            
        except Exception as e:
            self.progress_bar.setVisible(False)
            error_msg = f"Gagal memuat file {jenis}:\n{str(e)}"
            print(f"✗ {error_msg}")
            QMessageBox.critical(self, "Error", error_msg)
            
    def refresh_data(self):
        """Refresh current data untuk semua view"""
        if self.current_view_mode == 0:  # Konsinyasi
            if self.konsinyasi_df is not None:
                self.display_data_in_table(self.table_konsinyasi, self.konsinyasi_df, "KONSINYASI")
                self.status_bar.showMessage("✓ Data konsinyasi direfresh")
        elif self.current_view_mode == 1:  # Penjualan
            if self.penjualan_df is not None:
                self.display_data_in_table(self.table_penjualan, self.penjualan_df, "PENJUALAN")
                self.status_bar.showMessage("✓ Data penjualan direfresh")
        elif self.current_view_mode == 2:  # Gabungan
            self.update_gabungan_view()  # ✅ GUNAKAN METODE BARU
            self.status_bar.showMessage("✓ Data gabungan direfresh")
                
    def process_fallback(self, df, jenis):
        """Fallback processing jika LPDProcessor tidak tersedia - DIPERBAIKI"""
        if df is None or df.empty:
            print(f"⚠ Data {jenis} kosong di fallback")
            return pd.DataFrame()
        
        print(f"\n⚠ MENGGUNAKAN FALLBACK PROCESSING UNTUK {jenis}")
        
        try:
            # Lakukan processing sederhana berdasarkan format
            processed_df = df.copy()
            
            # Deteksi format manual
            column_names = [str(col).upper().strip() for col in processed_df.columns]
            
            # Coba cari kolom yang relevan
            kd_item_col = None
            nama_col = None
            jml_col = None
            
            for col in processed_df.columns:
                col_str = str(col).upper()
                
                if kd_item_col is None and any(word in col_str for word in ['KD', 'ITEM', 'KODE']):
                    kd_item_col = col
                if nama_col is None and any(word in col_str for word in ['NAMA', 'PELANGGAN']):
                    nama_col = col
                if jml_col is None and any(word in col_str for word in ['JML', 'JUMLAH', 'QTY']):
                    jml_col = col
            
            print(f"  - Kolom ditemukan: Kd.Item={kd_item_col}, Nama={nama_col}, Jml={jml_col}")
            
            # Deteksi cabang dari nama file
            if jenis == "konsinyasi" and self.konsinyasi_file:
                cabang_nama, kode_cabang = self.detect_cabang_from_filename(os.path.basename(self.konsinyasi_file))
            elif jenis == "penjualan" and self.penjualan_file:
                cabang_nama, kode_cabang = self.detect_cabang_from_filename(os.path.basename(self.penjualan_file))
            else:
                cabang_nama, kode_cabang = "JAKARTA", "GJK"
            
            # Tambahkan kolom informasi
            processed_df['JENIS'] = jenis
            processed_df['CABANG'] = cabang_nama
            processed_df['KODE_CABANG'] = kode_cabang
            
            # Jika ada kolom kd_item dan nama, buat summary sederhana
            if kd_item_col and nama_col and jml_col:
                try:
                    # Group by customer
                    summary = processed_df.groupby([nama_col, kd_item_col])[jml_col].sum().reset_index()
                    print(f"  - Summary dibuat: {len(summary)} baris")
                    
                    # Pivot untuk format template sederhana
                    pivot_df = summary.pivot_table(
                        index=nama_col,
                        columns=kd_item_col,
                        values=jml_col,
                        aggfunc='sum',
                        fill_value=0
                    ).reset_index()
                    
                    # Tambahkan total
                    pivot_df['TOTAL_PENJUALAN'] = pivot_df.select_dtypes(include=[np.number]).sum(axis=1)
                    
                    processed_df = pivot_df
                    
                except Exception as e:
                    print(f"  ⚠ Gagal buat summary: {e}")
            
            print(f"✓ Fallback {jenis} selesai: {len(processed_df)} baris")
            return processed_df
            
        except Exception as e:
            print(f"✗ Error di fallback processing: {e}")
            return df.copy()  # Return original jika error

    
    def get_sku_columns_from_df(self, df):
        """Dapatkan kolom SKU dari DataFrame"""
        if df is None or df.empty:
            return []
        
        # Cari kolom yang merupakan SKU
        sku_columns = []
        for col in df.columns:
            col_str = str(col).strip().upper()
            
            # Skip kolom yang jelas bukan SKU
            if any(keyword in col_str for keyword in [
                'NAMA', 'PELANGGAN', 'CUSTOMER', 'CUST', 
                'TOTAL', 'JUMLAH', 'JML', 'QTY', 
                'TANGGAL', 'DATE', 'TGL',
                'KODE', 'KD', 'ITEM', 'BARANG',
                'SISTEM', 'JENIS', 'CABANG', 'BRANCH',
                'PERINGKAT', 'RANK', 'PENCAPAIAN'
            ]):
                continue
                
            # Cek apakah ini SKU default
            if col_str in [sku.upper() for sku in self.default_skus]:
                sku_columns.append(col)
            # Cek apakah ini SKU dinamis
            elif col_str in [sku.upper() for sku in self.all_dynamic_skus]:
                sku_columns.append(col)
            # Cek pattern SKU (2-6 karakter, huruf/angka kombinasi)
            elif len(col_str) >= 2 and len(col_str) <= 6:
                # Pattern: minimal 1 huruf, tidak semua angka
                if any(c.isalpha() for c in col_str) and not col_str.isdigit():
                    # Cek apakah kolom ini berisi data numeric atau tidak
                    try:
                        # Sample beberapa data untuk menentukan tipe
                        sample_size = min(10, len(df))
                        sample_values = df[col].head(sample_size)
                        
                        # Jika sample mengandung string, skip
                        if sample_values.apply(lambda x: isinstance(x, str) and not x.replace('.', '', 1).isdigit()).any():
                            continue
                            
                        sku_columns.append(col)
                    except:
                        continue
        
        return sku_columns

    def process_file(self, jenis):
        """Proses file sesuai jenis - DIPERBAIKI untuk LPDProcessor baru"""
        try:
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(10)
            
            print(f"\n{'='*60}")
            print(f"PROSES FILE {jenis.upper()}")
            print(f"{'='*60}")
            
            if jenis == "konsinyasi":
                if self.konsinyasi_df is None:
                    QMessageBox.warning(self, "Peringatan", "Tidak ada data konsinyasi untuk diproses!")
                    return
                    
                # Proses dengan LPDProcessor jika tersedia
                if LPD_AVAILABLE:
                    processor = LPDProcessor()
                    print(f"✓ Memproses file: {os.path.basename(self.konsinyasi_file)}")
                    processed_df = processor.process_excel_to_template(self.konsinyasi_file)
                    self.konsinyasi_processed = processed_df.copy()
                    self.konsinyasi_df = processed_df
                    
                    # TAMBAHKAN PERHITUNGAN SKU UNTUK KONSINYASI
                    self.calculate_sku_statistics_for_df(self.konsinyasi_processed, "KONSINYASI")
                    
                    # Ambil SKU dinamis dari processor
                    if hasattr(processor, 'dynamic_skus'):
                        self.konsinyasi_skus = processor.dynamic_skus
                        print(f"✓ SKU dinamis terdeteksi dari konsinyasi: {len(self.konsinyasi_skus)} SKU")
                        
                        # Update all_dynamic_skus
                        for sku in self.konsinyasi_skus:
                            if sku not in self.all_dynamic_skus and sku not in self.default_skus:
                                self.all_dynamic_skus.append(sku)
                    
                    # Update all_skus
                    self.all_skus = self.default_skus + self.all_dynamic_skus
                    
                else:
                    # Fallback processing
                    processed_df = self.process_fallback(self.konsinyasi_df, "KONSINYASI")
                    self.konsinyasi_processed = processed_df
                    # TAMBAHKAN PERHITUNGAN SKU UNTUK FALLBACK
                    self.calculate_sku_statistics_for_df(self.konsinyasi_processed, "KONSINYASI")
                    
            elif jenis == "penjualan":
                if self.penjualan_df is None:
                    QMessageBox.warning(self, "Peringatan", "Tidak ada data penjualan untuk diproses!")
                    return
                    
                if LPD_AVAILABLE:
                    processor = LPDProcessor()
                    print(f"✓ Memproses file: {os.path.basename(self.penjualan_file)}")
                    processed_df = processor.process_excel_to_template(self.penjualan_file)
                    self.penjualan_processed = processed_df.copy()
                    self.penjualan_df = processed_df
                    
                    # TAMBAHKAN PERHITUNGAN SKU UNTUK PENJUALAN
                    self.calculate_sku_statistics_for_df(self.penjualan_processed, "PENJUALAN")
                    
                    # Ambil SKU dinamis dari processor
                    if hasattr(processor, 'dynamic_skus'):
                        self.penjualan_skus = processor.dynamic_skus
                        print(f"✓ SKU dinamis terdeteksi dari penjualan: {len(self.penjualan_skus)} SKU")
                        
                        # Update all_dynamic_skus
                        for sku in self.penjualan_skus:
                            if sku not in self.all_dynamic_skus and sku not in self.default_skus:
                                self.all_dynamic_skus.append(sku)
                    
                    # Update all_skus
                    self.all_skus = self.default_skus + self.all_dynamic_skus
                    
                else:
                    print("⚠ LPDProcessor tidak tersedia, gunakan fallback")
                    processed_df = self.process_fallback(self.penjualan_df, "PENJUALAN")
                    self.penjualan_processed = processed_df
                    # TAMBAHKAN PERHITUNGAN SKU UNTUK FALLBACK
                    self.calculate_sku_statistics_for_df(self.penjualan_processed, "PENJUALAN")
            
            self.progress_bar.setValue(100)
            QTimer.singleShot(500, lambda: self.progress_bar.setVisible(False))
            
            # Update view
            self.update_current_view()
            self.export_btn.setEnabled(True)
            
            self.status_bar.showMessage(f"✓ Data {jenis} berhasil diproses")
            
            # Tampilkan summary di stats label dengan info SKU
            if jenis == "konsinyasi" and self.konsinyasi_processed is not None:
                # DAPATKAN STATISTIK SKU
                stats = self.calculate_sku_statistics_for_df(self.konsinyasi_processed, "KONSINYASI", display=False)
                sku_count = stats.get('total_sku_ditemukan', 0)
                self.stats_label.setText(
                    f"📋 KONSINYASI: {len(self.konsinyasi_processed)} baris, "
                    f"{sku_count} SKU ditemukan, "
                    f"{stats.get('persen_masuk', 0):.1f}% ketercapaian"
                )
            elif jenis == "penjualan" and self.penjualan_processed is not None:
                # DAPATKAN STATISTIK SKU
                stats = self.calculate_sku_statistics_for_df(self.penjualan_processed, "PENJUALAN", display=False)
                sku_count = stats.get('total_sku_ditemukan', 0)
                self.stats_label.setText(
                    f"💰 PENJUALAN: {len(self.penjualan_processed)} baris, "
                    f"{sku_count} SKU ditemukan, "
                    f"{stats.get('persen_masuk', 0):.1f}% ketercapaian"
                )
            
        except Exception as e:
            self.progress_bar.setVisible(False)
            error_msg = f"Gagal memproses {jenis}: {str(e)}"
            print(f"✗ {error_msg}")
            import traceback
            traceback.print_exc()
            QMessageBox.warning(self, "Warning", error_msg)
        
    def update_current_view(self):
        """Update tampilan berdasarkan mode saat ini"""
        if self.current_view_mode == 0:  # Konsinyasi
            if self.konsinyasi_processed is not None:
                self.display_data_in_table(self.table_konsinyasi, self.konsinyasi_processed, "KONSINYASI")
                self.export_btn.setEnabled(True)
            elif self.konsinyasi_df is not None:
                self.display_data_in_table(self.table_konsinyasi, self.konsinyasi_df, "KONSINYASI RAW")
                self.export_btn.setEnabled(True)
                
        elif self.current_view_mode == 1:  # Penjualan
            if self.penjualan_processed is not None:
                self.display_data_in_table(self.table_penjualan, self.penjualan_processed, "PENJUALAN")
                self.export_btn.setEnabled(True)
            elif self.penjualan_df is not None:
                self.display_data_in_table(self.table_penjualan, self.penjualan_df, "PENJUALAN RAW")
                self.export_btn.setEnabled(True)
                
        elif self.current_view_mode == 2:  # Gabungan
            self.update_gabungan_view()  # Gunakan method baru
            has_data = (self.konsinyasi_processed is not None and not self.konsinyasi_processed.empty) or \
                    (self.penjualan_processed is not None and not self.penjualan_processed.empty) or \
                    (self.konsinyasi_df is not None and not self.konsinyasi_df.empty) or \
                    (self.penjualan_df is not None and not self.penjualan_df.empty)
            self.export_btn.setEnabled(has_data)

    def display_data_in_table(self, table_widget, df, jenis):
        """Display data di tabel tertentu dengan formatting yang optimal"""
        table_widget.setRowCount(0)
        table_widget.setColumnCount(0)
        
        if df is None or df.empty:
            return
        
        # Disable sorting sementara
        table_widget.setSortingEnabled(False)
        
        # Set tabel dengan font yang lebih besar
        table_widget.setRowCount(len(df))
        table_widget.setColumnCount(len(df.columns))
        table_widget.setHorizontalHeaderLabels(df.columns.tolist())
        
        # Font lebih besar untuk header
        font = QFont("Arial", 11, QFont.Bold)
        table_widget.horizontalHeader().setFont(font)
        table_widget.horizontalHeader().setDefaultSectionSize(100)
        
        # Get SKU columns for special formatting
        sku_columns = self.get_sku_columns_from_df(df)
        
        # TAMBAHKAN: Hitung statistik SKU untuk ditampilkan
        sku_stats = self.calculate_sku_statistics_for_df(df, jenis, display=False)
        
        print(f"📊 {jenis} - Statistik SKU:")
        print(f"  • Total SKU ditemukan: {sku_stats['total_sku_ditemukan']}")
        print(f"  • SKU masuk: {sku_stats['sku_masuk']} ({sku_stats['persen_masuk']:.1f}%)")
        print(f"  • SKU tidak masuk: {sku_stats['sku_tidak_masuk']} ({sku_stats['persen_tidak_masuk']:.1f}%)")
        
        
        # Isi data dengan formatting
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                cell_value = df.iat[row_idx, col_idx]
                col_name = str(df.columns[col_idx])
                
                # Handle NaN values
                if pd.isna(cell_value):
                    item = QTableWidgetItem("")
                else:
                    item = QTableWidgetItem(str(cell_value))
                
                # Font lebih besar untuk data
                item.setFont(QFont("Arial", 10))
                
                # Format numeric values
                if isinstance(cell_value, (int, float)) and not pd.isna(cell_value):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    if isinstance(cell_value, int) or (isinstance(cell_value, float) and cell_value.is_integer()):
                        item.setText(f"{int(cell_value):,}")
                    else:
                        item.setText(f"{cell_value:,.2f}")
                    
                    # Highlight nilai positif untuk SKU
                    if col_name in sku_columns and cell_value > 0:
                        item.setForeground(QColor(39, 174, 96))  # Green
                        item.setFont(QFont("Arial", 10, QFont.Bold))
                else:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                
                # Beri warna background berdasarkan jenis
                if jenis == "KONSINYASI":
                    if row_idx % 2 == 0:
                        item.setBackground(QColor(232, 244, 248))  # Biru muda
                    else:
                        item.setBackground(QColor(242, 250, 252))  # Biru lebih muda
                else:  # PENJUALAN
                    if row_idx % 2 == 0:
                        item.setBackground(QColor(248, 232, 244))  # Pink muda
                    else:
                        item.setBackground(QColor(252, 242, 250))  # Pink lebih muda
                
                # Highlight kolom SKU dengan bold dan warna
                if col_name in sku_columns:
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                    item.setForeground(QColor(155, 89, 182))  # Ungu untuk SKU
                
                # Highlight baris TOTAL dengan warna berbeda
                if isinstance(cell_value, str) and 'TOTAL' in str(cell_value).upper():
                    item.setBackground(QColor(220, 220, 220))
                    item.setFont(QFont("Arial", 11, QFont.Bold))
                    item.setForeground(QColor(0, 0, 0))
                
                table_widget.setItem(row_idx, col_idx, item)
        
        # Enable sorting
        table_widget.setSortingEnabled(True)
        
        # Auto resize columns
        table_widget.resizeColumnsToContents()
        
        # Set minimum column width
        for col in range(table_widget.columnCount()):
            width = table_widget.columnWidth(col)
            if width < 80:
                table_widget.setColumnWidth(col, 80)
        
        # Tampilkan info tentang jumlah SKU
        print(f"✓ {jenis} ditampilkan: {len(df)} baris, {len(df.columns)} kolom "
            f"({sku_stats['total_sku_ditemukan']} kolom SKU, "
            f"{sku_stats['sku_masuk']} SKU terjual)")
        
        # TAMBAHKAN: Update stats label dengan info SKU yang lebih detail
        if jenis == "KONSINYASI":
            self.stats_label.setText(
                f"📋 {jenis}: {len(df)} baris, "
                f"{sku_stats['total_sku_ditemukan']} SKU ditemukan, "
                f"{sku_stats['sku_masuk']} SKU terjual ({sku_stats['persen_masuk']:.1f}%)"
            )
        elif jenis == "PENJUALAN":
            self.stats_label.setText(
                f"💰 {jenis}: {len(df)} baris, "
                f"{sku_stats['total_sku_ditemukan']} SKU ditemukan, "
                f"{sku_stats['sku_masuk']} SKU terjual ({sku_stats['persen_masuk']:.1f}%)"
            )
                    
    def auto_resize_all_tables(self):
        """Auto resize semua tabel"""
        if hasattr(self, 'table_konsinyasi'):
            self.table_konsinyasi.resizeColumnsToContents()
        
        if hasattr(self, 'table_penjualan'):
            self.table_penjualan.resizeColumnsToContents()
        
        if hasattr(self, 'table_gabungan'):  # Hanya 1 tabel
            self.table_gabungan.resizeColumnsToContents()
        
        self.status_bar.showMessage("✓ Semua tabel di-resize")
            

    def export_data(self):
        # Tentukan data mana yang akan diexport berdasarkan view mode
        if self.current_view_mode == 0:  # Konsinyasi
            df = self.konsinyasi_df
            judul = "LAPORAN KONSINYASI"
            default_name = "Laporan_Konsinyasi"
        elif self.current_view_mode == 1:  # Penjualan
            df = self.penjualan_df
            judul = "LAPORAN PENJUALAN"
            default_name = "Laporan_Penjualan"
        elif self.current_view_mode == 2:  # Gabungan
            self.export_gabungan_data()
            return
        
        if df is None or df.empty:
            QMessageBox.warning(self, "Warning", "Tidak ada data untuk diexport!")
            return
        
        # Generate timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export ke Excel",
            f"{default_name}_{timestamp}.xlsx",
            "Excel Files (*.xlsx);;CSV Files (*.csv)"
        )
        
        if file_path:
            try:
                self.progress_bar.setVisible(True)
                self.progress_bar.setValue(30)
                
                if file_path.endswith('.csv'):
                    df.to_csv(file_path, index=False, encoding='utf-8')
                else:
                    # Simpan DataFrame ke Excel tanpa header sementara
                    df.to_excel(file_path, index=False)
                    
                    # Load workbook untuk formatting
                    wb = load_workbook(file_path)
                    ws = wb.active
                    
                    # **PERBAIKAN: Geser data ke bawah untuk memberi ruang header judul**
                    # Shift semua data 2 baris ke bawah
                    ws.insert_rows(1, 2)  # Sisipkan 2 baris kosong di atas
                    
                    # Merge header untuk judul
                    max_col = ws.max_column
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
                    cell = ws.cell(row=1, column=1)
                    cell.value = judul
                    cell.font = Font(bold=True, size=14)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[1].height = 32
                    
                    # **PERBAIKAN: Auto column width yang AMAN dari merged cells**
                    for col_idx, col in enumerate(ws.iter_cols(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 1):
                        max_length = 0
                        col_letter = None
                        
                        # Dapatkan huruf kolom dari cell pertama yang bukan merged
                        for cell in col:
                            if not cell.coordinate in ws.merged_cells:
                                try:
                                    # Dapatkan kolom letter dengan aman
                                    if hasattr(cell, 'column_letter'):
                                        col_letter = cell.column_letter
                                    else:
                                        # Alternatif: hitung dari index kolom
                                        col_letter = openpyxl.utils.get_column_letter(cell.column)
                                    break
                                except:
                                    continue
                        
                        if col_letter is None:
                            # Fallback: hitung dari index
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                        
                        # Cari panjang maksimum dalam kolom ini (hanya data, bukan header)
                        for cell in col:
                            try:
                                if cell.value:
                                    cell_length = len(str(cell.value))
                                    if cell_length > max_length:
                                        max_length = cell_length
                            except:
                                continue
                        
                        # Set width (minimum 8, maximum 50)
                        adjusted_width = min(max_length + 2, 50)
                        if adjusted_width < 8:
                            adjusted_width = 8
                        
                        ws.column_dimensions[col_letter].width = adjusted_width
                    
                    # Format header data (baris 3)
                    for cell in ws[3]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")
                    
                    wb.save(file_path)
                
                self.progress_bar.setValue(100)
                QTimer.singleShot(500, lambda: self.progress_bar.setVisible(False))
                
                QMessageBox.information(self, "Sukses", 
                    f"Data berhasil diexport ke:\n{file_path}\n\n"
                    f"Jumlah baris: {len(df)}\n"
                    f"Jumlah kolom: {len(df.columns)}")
                
            except Exception as e:
                self.progress_bar.setVisible(False)
                error_msg = f"Gagal mengexport:\n{str(e)}"
                print(f"❌ Error export: {str(e)}")
                import traceback
                traceback.print_exc()
                QMessageBox.critical(self, "Error", error_msg)
    
    def _extract_sku_count_from_row(self, row):
        """Extract jumlah SKU yang dibeli dari sebuah baris"""
        try:
            # Asumsi: kolom SKU berada di posisi 4 sampai sebelum kolom statistik SKU
            # Kolom statistik SKU dimulai dari "SKU YANG SUDAH MASUK"
            
            # Cari index kolom "SKU YANG SUDAH MASUK"
            # Ini adalah pendekatan sederhana
            if len(row) < 10:
                return 0
                
            # Hitung SKU dengan nilai > 0
            sku_count = 0
            # Asumsi: kolom 4-32 adalah kolom SKU (28 SKU)
            for i in range(4, min(32, len(row))):
                value = row[i]
                if self._extract_number_from_value(value) > 0:
                    sku_count += 1
                    
            return sku_count
        except:
            return 0
    
    def export_gabungan_data(self):
        """Export data gabungan ke Excel PERSIS seperti view gabungan"""
        try:
            # Cek apakah sedang di view gabungan
            if self.current_view_mode != 2:
                QMessageBox.warning(self, "Peringatan", 
                    "Anda harus berada di VIEW GABUNGAN untuk export data gabungan!")
                return
            
            # Cek apakah tabel gabungan memiliki data
            if self.table_gabungan.rowCount() == 0 or self.table_gabungan.columnCount() == 0:
                QMessageBox.warning(self, "Peringatan", 
                    "Tabel gabungan kosong! Tidak ada data untuk diexport.")
                return
            
            # Generate timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Data Gabungan ke Excel",
                f"Laporan_Gabungan_{timestamp}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
                
            try:
                self.progress_bar.setVisible(True)
                self.progress_bar.setValue(10)
                
                print(f"\n{'='*60}")
                print(f"📤 EXPORT DATA GABUNGAN PERSIS DENGAN VIEW")
                print(f"{'='*60}")
                
                # ====================== AMBIL DATA DARI TABLE GABUNGAN ======================
                print("📋 Mengambil data dari tabel gabungan...")
                
                # Dapatkan header dari tabel
                headers = []
                for col in range(self.table_gabungan.columnCount()):
                    header_item = self.table_gabungan.horizontalHeaderItem(col)
                    if header_item:
                        headers.append(header_item.text())
                    else:
                        headers.append(f"Kolom_{col+1}")
                
                # Dapatkan semua data dari tabel
                all_data = []
                for row in range(self.table_gabungan.rowCount()):
                    row_data = []
                    for col in range(self.table_gabungan.columnCount()):
                        item = self.table_gabungan.item(row, col)
                        if item:
                            cell_text = item.text()
                            
                            # Coba konversi ke angka jika memungkinkan
                            try:
                                # Hapus pemisah ribuan, tanda %, dan karakter non-digit
                                clean_text = cell_text.replace(',', '').replace('.', '').replace('%', '')
                                if clean_text.isdigit():
                                    # Format ribuan atau desimal
                                    if ',' in cell_text:
                                        cell_value = int(clean_text)
                                    elif '.' in cell_text and cell_text.count('.') == 1:
                                        # Cek apakah ini angka desimal
                                        parts = cell_text.replace(',', '').split('.')
                                        if parts[0].isdigit() and parts[1].isdigit():
                                            cell_value = float(cell_text.replace(',', ''))
                                        else:
                                            cell_value = cell_text
                                    else:
                                        cell_value = float(cell_text) if '.' in cell_text else int(cell_text)
                                else:
                                    # Cek apakah ini persentase
                                    if cell_text.endswith('%'):
                                        try:
                                            cell_value = float(cell_text.replace('%', '')) / 100
                                        except:
                                            cell_value = cell_text
                                    else:
                                        cell_value = cell_text
                            except:
                                cell_value = cell_text
                            
                            row_data.append(cell_value)
                        else:
                            row_data.append("")
                    
                    all_data.append(row_data)
                
                print(f"✓ Data diambil: {len(all_data)} baris, {len(headers)} kolom")
                
                self.progress_bar.setValue(30)
                
                # ====================== BUAT DATAFRAME DARI DATA TABEL ======================
                df_from_table = pd.DataFrame(all_data, columns=headers)
                
                print(f"✓ DataFrame dibuat: {df_from_table.shape}")
                
                # ====================== BUAT WORKBOOK EXCEL ======================
                print("📊 Membuat workbook Excel...")
                
                # Buat workbook baru tanpa writer
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "LAPORAN_GABUNGAN"
                
                self.progress_bar.setValue(40)
                
                # ====================== TAMBAHKAN HEADER DAN JUDUL ======================
                print("✏️ Menambahkan header dan judul...")
                
                # 1. Judul utama
                title_cell = ws.cell(row=1, column=1)
                title_cell.value = "LAPORAN GABUNGAN KONSINYASI & PENJUALAN"
                title_cell.font = Font(bold=True, size=16, color="000000")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Merge cells untuk judul (HANYA JIKA ADA LEBIH DARI 1 KOLOM)
                if len(headers) > 1:
                    try:
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
                    except Exception as e:
                        print(f"⚠ Gagal merge cells untuk judul: {e}")
                        # Tetap lanjut tanpa merge
                
                ws.row_dimensions[1].height = 30
                
                # 2. Subjudul (tanggal dan info)
                subtitle_cell = ws.cell(row=2, column=1)
                export_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                subtitle_cell.value = f"Export tanggal: {export_time} | Sumber: Excel Report Processor v1.0"
                subtitle_cell.font = Font(size=10, italic=True, color="666666")
                subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if len(headers) > 1:
                    try:
                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
                    except Exception as e:
                        print(f"⚠ Gagal merge cells untuk subjudul: {e}")
                
                ws.row_dimensions[2].height = 20
                
                # 3. Baris kosong sebagai pemisah
                ws.row_dimensions[3].height = 10
                
                self.progress_bar.setValue(50)
                
                # ====================== TULIS HEADER TABEL ======================
                print("🎨 Menulis header tabel...")
                
                # Header tabel (mulai dari row 4)
                header_row = 4
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=header_row, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    # Warna header berdasarkan jenis
                    header_text = str(header).upper()
                    
                    # Kolom JENIS
                    if header_text == "JENIS":
                        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                    # Kolom dasar (NAMA, SISTEM, TOTAL)
                    elif any(keyword in header_text for keyword in ['NAMA', 'PELANGGAN', 'SISTEM', 'TOTAL']):
                        cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                    # Kolom SKU
                    elif (header_text in [sku.upper() for sku in self.default_skus] or 
                        header_text in [sku.upper() for sku in self.all_dynamic_skus] or
                        self._is_pure_sku_column(header)):
                        cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
                    # Kolom statistik SKU
                    elif self._is_sku_stat_column(header):
                        cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                    # Default
                    else:
                        cell.fill = PatternFill(start_color="7F8C8D", end_color="7F8C8D", fill_type="solid")
                    
                    # Border untuk header
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='medium', color='000000')
                    )
                
                ws.row_dimensions[header_row].height = 25
                
                self.progress_bar.setValue(60)
                
                # ====================== TULIS DATA ======================
                print("📝 Menulis data ke Excel...")
                
                total_rows = len(all_data)
                
                # Tulis data baris per baris
                for row_idx, row_data in enumerate(all_data):
                    excel_row = header_row + 1 + row_idx
                    
                    # Pastikan row_data memiliki panjang yang sama dengan headers
                    if len(row_data) < len(headers):
                        # Tambahkan nilai kosong jika kurang
                        row_data = row_data + [""] * (len(headers) - len(row_data))
                    elif len(row_data) > len(headers):
                        # Potong jika terlalu panjang
                        row_data = row_data[:len(headers)]
                    
                    # Cek jenis data dari kolom pertama
                    jenis = ""
                    if len(row_data) > 0:
                        jenis = str(row_data[0]).upper()
                    
                    # Tulis data untuk setiap kolom
                    for col_idx, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=excel_row, column=col_idx)
                        
                        # Format nilai
                        if isinstance(cell_value, (int, float)):
                            # Format angka
                            cell.value = cell_value
                            
                            # Format khusus untuk persentase (dari nilai 0-1)
                            if 0 <= cell_value <= 1 and col_idx < len(headers):
                                col_name = headers[col_idx-1] if col_idx-1 < len(headers) else ""
                                if '%' in str(col_name):
                                    cell.number_format = '0.00%'
                                else:
                                    cell.number_format = '#,##0' if cell_value == int(cell_value) else '#,##0.00'
                            else:
                                cell.number_format = '#,##0' if isinstance(cell_value, int) else '#,##0.00'
                            
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            # Format teks
                            cell.value = str(cell_value) if cell_value is not None else ""
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        
                        # Warna latar berdasarkan jenis - DIPERBAIKI UNTUK UKURAN SEL
                        if jenis == "KONSINYASI":
                            if row_idx % 2 == 0:
                                cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="F2FAFC", end_color="F2FAFC", fill_type="solid")
                        
                        elif jenis == "PENJUALAN":
                            if row_idx % 2 == 0:
                                cell.fill = PatternFill(start_color="F8E8F4", end_color="F8E8F4", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FCF2FA", end_color="FCF2FA", fill_type="solid")
                        
                        elif "TOTAL KONSINYASI" in jenis:
                            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                            if isinstance(cell_value, (int, float, str)):
                                cell.font = Font(bold=True)
                        
                        elif "TOTAL PENJUALAN" in jenis:
                            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                            if isinstance(cell_value, (int, float, str)):
                                cell.font = Font(bold=True)
                        
                        elif "TOTAL GABUNGAN" in jenis:
                            cell.fill = PatternFill(start_color="FFFFC8", end_color="FFFFC8", fill_type="solid")
                            if isinstance(cell_value, (int, float, str)):
                                cell.font = Font(bold=True, size=11)
                        
                        # Border untuk sel data
                        cell.border = Border(
                            left=Side(style='thin', color='DDDDDD'),
                            right=Side(style='thin', color='DDDDDD'),
                            top=Side(style='thin', color='DDDDDD'),
                            bottom=Side(style='thin', color='DDDDDD')
                        )
                    
                    # Atur tinggi baris
                    ws.row_dimensions[excel_row].height = 20
                
                self.progress_bar.setValue(80)
                
                # ====================== SESUAIKAN LEBAR KOLOM ======================
                print("📏 Menyesuaikan lebar kolom...")
                
                # Atur lebar kolom yang optimal
                column_widths = {}
                
                for col_idx in range(1, len(headers) + 1):
                    max_length = 0
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    
                    # Cek panjang di header
                    header_text = str(headers[col_idx-1]) if col_idx-1 < len(headers) else ""
                    if header_text:
                        max_length = len(header_text)
                    
                    # Cek panjang di data
                    for row_idx in range(header_row + 1, header_row + 1 + total_rows):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    
                    # Tentukan lebar optimal
                    optimal_width = min(max_length + 2, 40)  # Maksimum 40 karakter
                    if optimal_width < 8:  # Minimum 8 karakter
                        optimal_width = 8
                    
                    # Lebar khusus untuk kolom tertentu
                    header_text_upper = header_text.upper()
                    if self._is_sku_stat_column(header_text):
                        if 'SKU YANG' in header_text_upper:
                            optimal_width = 25
                        elif 'JML SKU' in header_text_upper or 'JUMLAH SKU' in header_text_upper:
                            optimal_width = 12
                        elif header_text_upper == '%':
                            optimal_width = 8
                        elif 'PERINGKAT' in header_text_upper:
                            optimal_width = 15
                    elif 'JENIS' in header_text_upper:
                        optimal_width = 12
                    elif 'NAMA' in header_text_upper or 'PELANGGAN' in header_text_upper:
                        optimal_width = 30
                    elif 'SISTEM' in header_text_upper:
                        optimal_width = 12
                    elif 'TOTAL' in header_text_upper:
                        optimal_width = 15
                    elif header_text_upper in [sku.upper() for sku in self.default_skus + self.all_dynamic_skus]:
                        optimal_width = 10  # Lebar standar untuk kolom SKU
                    
                    ws.column_dimensions[col_letter].width = optimal_width
                
                self.progress_bar.setValue(90)
                
                # ====================== TAMBAHKAN FITUR EXCEL ======================
                print("🔧 Menambahkan fitur Excel...")
                
                # Freeze panes (header tetap terlihat)
                try:
                    ws.freeze_panes = f"A{header_row + 1}"
                except:
                    pass
                
                # Auto filter
                try:
                    filter_range = f"A{header_row}:{openpyxl.utils.get_column_letter(len(headers))}{header_row + total_rows}"
                    ws.auto_filter.ref = filter_range
                except:
                    pass
                
                # ====================== SIMPAN FILE ======================
                print("💾 Menyimpan file Excel...")
                wb.save(file_path)
                
                self.progress_bar.setValue(100)
                
                print(f"✅ Excel berhasil dibuat: {file_path}")
                print(f"   • Total baris: {total_rows}")
                print(f"   • Total kolom: {len(headers)}")
                print(f"   • Sheet: LAPORAN_GABUNGAN")
                
                QTimer.singleShot(500, lambda: self.progress_bar.setVisible(False))
                
                # ====================== TAMPILKAN KONFIRMASI ======================
                # Hitung statistik
                kons_count = sum(1 for row in all_data if len(row) > 0 and str(row[0]).upper() == "KONSINYASI")
                penj_count = sum(1 for row in all_data if len(row) > 0 and str(row[0]).upper() == "PENJUALAN")
                total_count = sum(1 for row in all_data if len(row) > 0 and "TOTAL" in str(row[0]).upper())
                detail_count = total_rows - (kons_count + penj_count + total_count)
                
                # PERBAIKAN: Cara yang benar untuk membuat custom QMessageBox
                success_msg = QMessageBox()
                success_msg.setWindowTitle("✅ Export Berhasil")
                success_msg.setIcon(QMessageBox.Information)
                
                message = f"""
                <h3>Export Berhasil!</h3>
                <p>Data gabungan berhasil diexport ke:</p>
                <p style="font-weight: bold; color: #2C3E50;">{os.path.basename(file_path)}</p>
                
                <table border='0' cellpadding='3' style="margin: 10px 0;">
                <tr><td colspan='2'><b>📊 Statistik Export:</b></td></tr>
                <tr><td>Total baris:</td><td style='text-align: right; font-weight: bold;'>{total_rows}</td></tr>
                <tr><td>Total kolom:</td><td style='text-align: right; font-weight: bold;'>{len(headers)}</td></tr>
                <tr><td>KONSINYASI:</td><td style='text-align: right; font-weight: bold;'>{kons_count}</td></tr>
                <tr><td>PENJUALAN:</td><td style='text-align: right; font-weight: bold;'>{penj_count}</td></tr>
                <tr><td>BARIS TOTAL:</td><td style='text-align: right; font-weight: bold;'>{total_count}</td></tr>
                <tr><td>Detail:</td><td style='text-align: right; font-weight: bold;'>{detail_count}</td></tr>
                </table>
                
                <p><i>File Excel berisi format yang sama persis dengan tampilan di aplikasi.</i></p>
                <p><small>Lokasi: {file_path}</small></p>
                """
                
                success_msg.setTextFormat(Qt.RichText)
                success_msg.setText(message)
                
                # PERBAIKAN: Buat tombol custom dengan cara yang benar
                open_button = QPushButton("📂 Buka File")
                ok_button = QPushButton("OK")
                
                success_msg.addButton(ok_button, QMessageBox.AcceptRole)
                success_msg.addButton(open_button, QMessageBox.ActionRole)
                
                result = success_msg.exec_()
                
                # Buka file jika tombol "Buka File" diklik
                if result == QMessageBox.ActionRole or success_msg.clickedButton() == open_button:
                    try:
                        import subprocess
                        if sys.platform == "win32":
                            os.startfile(file_path)
                        elif sys.platform == "darwin":  # macOS
                            subprocess.call(['open', file_path])
                        else:  # linux
                            subprocess.call(['xdg-open', file_path])
                    except Exception as e:
                        print(f"⚠ Tidak bisa membuka file: {e}")
                        QMessageBox.information(self, "Info", f"File berhasil diexport ke:\n{file_path}")
                
            except Exception as e:
                self.progress_bar.setVisible(False)
                error_msg = f"Gagal mengexport data gabungan:\n\n{str(e)}"
                print(f"❌ Error export: {str(e)}")
                import traceback
                traceback.print_exc()
                QMessageBox.critical(self, "❌ Error", error_msg)
                
        except Exception as e:
            self.progress_bar.setVisible(False)
            error_msg = f"Error dalam proses export:\n\n{str(e)}"
            QMessageBox.critical(self, "❌ Error", error_msg)
                            
    def clear_table_selection(self):
        """Clear table selection untuk semua tabel"""
        tables = [
            self.table_konsinyasi,
            self.table_penjualan,
        ]
        
        for table in tables:
            if table is not None:
                table.clearSelection()
                    
    def show_about(self):
        """Show about dialog"""
        about_text = """
        <h3>Excel Report Processor - Konsinyasi & Penjualan</h3>
        <p><b>Versi 1.0</b></p>
        <p>Aplikasi untuk memproses data konsinyasi/penjualan menjadi format laporan template</p>
        
        <p><b>Fitur Utama:</b></p>
        <ul>
        <li>📂 Baca file Excel/CSV data konsinyasi/penjualan</li>
        <li>⚙️ Proses otomatis ke format template laporan</li>
        <li>👁️ Tampilkan data input asli dan template</li>
        <li>📊 Format tabel dengan warna dan styling</li>
        <li>💾 Export laporan ke Excel/CSV</li>
        <li>🔍 Auto-detect kolom data</li>
        <li>🔄 Multi-sheet support</li>
        </ul>
        
        <p><b>Format Template:</b></p>
        <ul>
        <li>NAMA CUSTOMER</li>
        <li>TOTAL PENJUALAN</li>
        <li>JUDUL CABANG</li>
        <li>22 SKU (TO, TB, TBL, TS, dll.)</li>
        <li>PENCAPAIAN KODE SKU</li>
        </ul>
        
        <p><b>Penggunaan:</b></p>
        <ol>
        <li>Pilih file data melalui menu File atau tombol 📂</li>
        <li>Klik ⚙️ Proses ke Template untuk konversi</li>
        <li>Ganti mode tampilan antara Data Input dan Template</li>
        <li>Export hasil laporan dengan 💾 Export Laporan</li>
        </ol>
        
        <p><i>Developed for konsinyasi & penjualan reporting</i></p>
        """
        
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Tentang Aplikasi")
        msg_box.setTextFormat(Qt.RichText)
        msg_box.setText(about_text)
        msg_box.setIcon(QMessageBox.Information)
        msg_box.exec_()
    
    def calculate_sku_ranking_for_all_customers(self, df, ordered_skus, jenis):
        """Hitung peringkat SKU untuk semua pelanggan - URUTKAN SEMUA SKU 1-N"""
        try:
            if df is None or df.empty:
                return {}
            
            print(f"\n🏆 MENGHITUNG PERINGKAT SKU UNTUK {jenis} (SEMUA SKU)")

            # Buat dictionary untuk menyimpan peringkat per pelanggan
            ranking_dict = {}
            
            # Counter untuk statistik
            total_pelanggan = 0
            pelanggan_dengan_ranking = 0
            
            # Hitung untuk setiap baris (pelanggan)
            for idx, row in df.iterrows():
                # Dapatkan nama pelanggan (kolom pertama yang mengandung nama)
                nama_pelanggan = "Pelanggan"
                for col in df.columns:
                    if any(keyword in str(col).upper() for keyword in ['NAMA', 'PELANGGAN', 'CUSTOMER']):
                        nama_pelanggan = str(row[col])
                        break
                
                # Skip jika ini baris TOTAL
                if 'TOTAL' in str(nama_pelanggan).upper():
                    continue
                
                total_pelanggan += 1
                
                # Hitung total penjualan per SKU untuk pelanggan ini
                sku_totals = []
                
                for sku in ordered_skus:
                    if sku in df.columns:
                        value = row[sku]
                        qty = self._extract_number_from_value(value)
                        if qty > 0:
                            sku_totals.append({
                                'sku': sku,
                                'qty': qty
                            })
                
                # URUTKAN SEMUA SKU berdasarkan quantity (descending)
                sku_totals_sorted = sorted(sku_totals, key=lambda x: x['qty'], reverse=True)
                
                # Ambil SEMUA SKU yang dibeli (bukan hanya top 3)
                all_skus_ranking = []
                
                # Jika ada SKU yang dibeli
                if sku_totals_sorted:
                    pelanggan_dengan_ranking += 1
                    
                    # Format: "1. TO (50 pcs) | 2. TB (30 pcs) | 3. TWS (20 pcs) ..."
                    for i, item in enumerate(sku_totals_sorted):
                        all_skus_ranking.append(f"{i+1}. {item['sku']} ({item['qty']:.0f} pcs)")
                    
                    # Batasi maksimal 10 SKU dalam tampilan untuk readability
                    # Tapi data asli tetap simpan semua
                    if len(all_skus_ranking) > 10:
                        display_ranking = " | ".join(all_skus_ranking[:10]) + f" | ... ({len(all_skus_ranking)-10} SKU lainnya)"
                    else:
                        display_ranking = " | ".join(all_skus_ranking)
                else:
                    display_ranking = "Tidak ada pembelian"
                
                # Gabungkan menjadi string
                ranking_str = display_ranking
                
                # Simpan di dictionary - SIMPAN DATA LENGKAP JUGA
                key = f"{jenis}_{idx}"
                ranking_dict[key] = {
                    'nama': nama_pelanggan,
                    'ranking': ranking_str,
                    'all_skus': [item['sku'] for item in sku_totals_sorted],  # Semua SKU
                    'all_quantities': [item['qty'] for item in sku_totals_sorted],  # Semua quantity
                    'total_sku_terjual': len(sku_totals),
                    'total_qty': sum(item['qty'] for item in sku_totals),
                    'ranking_full': all_skus_ranking,  # Full ranking untuk ekspor
                    'ranking_score': 1000 - len(sku_totals_sorted) if sku_totals_sorted else -1,  # Skor untuk sorting
                    'jenis': jenis  # Tambahkan jenis untuk sorting
                }
            
            print(f"  ✓ Total pelanggan: {total_pelanggan}")
            print(f"  ✓ Pelanggan dengan ranking: {pelanggan_dengan_ranking}")
            
            # Tampilkan contoh ranking untuk debugging
            if ranking_dict:
                first_key = list(ranking_dict.keys())[0]
                first_data = ranking_dict[first_key]
                print(f"  ✓ Contoh ranking untuk {first_data['nama'][:20]}...:")
                print(f"    - {first_data['ranking'][:100]}...")
            
            return ranking_dict
            
        except Exception as e:
            print(f"❌ Error calculating SKU ranking: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
def main():
    app = QApplication(sys.argv)
    
    # Set application style
    app.setStyle("Fusion")
    
    # Set application palette
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor(240, 240, 240))
    palette.setColor(QPalette.WindowText, QColor(0, 0, 0))
    app.setPalette(palette)
    
    # Create and show main window
    window = ExcelReportProcessor()
    window.show()
    
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()

